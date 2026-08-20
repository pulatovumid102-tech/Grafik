"""
Sirly Mini-App uchun Telegram bot.

Vazifasi:
  - index.html (mini-app) Supabase'ga yozib qo'yadigan xabar/eslatma/fayl
    so'rovlarini muntazam tekshirib, Telegram orqali yetkazib beradi.
  - Uchta jadval kuzatiladi:
      1) bot_group_messages  -> umumiy guruhga (GROUP_CHAT_ID) xabar
      2) bot_reminders       -> aniq foydalanuvchiga (chat_id) belgilangan
                                 vaqtda eslatma (sekretar, kaiten budilnik)
      3) file_send_requests  -> Hujjatlar bo'limidan foydalanuvchiga
                                 xabar yoki fayl yuborish

Ishga tushirish:
  1) pip install -r requirements.txt
  2) .env.example'ni .env qilib nusxalab, qiymatlarni to'ldiring
  3) python bot.py
"""

import os
import logging
import io
from datetime import datetime, timezone, time as dt_time, timedelta
from zoneinfo import ZoneInfo

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, MessageHandler, filters

# ============================================================
# SOZLAMALAR — to'g'ridan-to'g'ri shu yerga yozilgan
# (Railway'da Variables qo'shish shart emas, lekin xohlasangiz
#  pastdagi os.environ.get(...) orqali Variables bilan ustidan
#  yozib qo'yish ham mumkin — hozircha shart emas)
# ============================================================
BOT_TOKEN = os.environ.get("BOT_TOKEN", "8693834890:AAHMcQyOmv2bP0xKmvHYl32jvst2BU4PgtY")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://qtrniovpkrwimeohamkc.supabase.co").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InF0cm5pb3Zwa3J3aW1lb2hhbWtjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMTY4NjUsImV4cCI6MjA5NjU5Mjg2NX0.j4gUqZlqMHR0ltIMCDB-UfWPvuPVs9B9HF0If2fPxhU")
GROUP_CHAT_ID = int(os.environ.get("GROUP_CHAT_ID", "-1003823489442"))
SUPPORT_GROUP_ID = -1003823489442
PARTNERSHIP_GROUP_ID = -5467968653
ZVONOK2_GROUP_ID = -5450377196
SIRLY_STAFF_GROUP_ID = -5076135815
MUAMMOLI_HAMKORLAR_REPORT_GROUP_ID = -1004336078331
HR_GROUP_ID = -5370864546
TOPSHIRIQLAR_GROUP_ID = -5550614907
BUXGALTERIYA_PROMOKOD_GROUP_ID = -5574268734
BAZA415_REPORT_GROUP_ID = -5228403271
MINIAPP_URL = os.environ.get("MINIAPP_URL", "https://pulatovumid102-tech.github.io/Grafik/")
POLL_SECONDS = int(os.environ.get("POLL_SECONDS", "15"))
BATCH_LIMIT = int(os.environ.get("BATCH_LIMIT", "20"))
MUAMMOLI_MUDDAT_SOAT = 24
TASHKENT_TZ = ZoneInfo("Asia/Tashkent")
KAITEN_DEPTS = [
    "Sotuv bo'limi",
    "Partnership bo'limi",
    "Support bo'limi",
    "IT bo'limi",
    "Buxgalteriya bo'limi",
    "Yuridik bo'limi",
    "Task Management",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
log = logging.getLogger("sirly-bot")

SB_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

http_client: httpx.AsyncClient | None = None


# ============================================================
# SUPABASE YORDAMCHI FUNKSIYALAR
# ============================================================
async def sb_get(table: str, params: dict) -> list:
    resp = await http_client.get(
        f"{SUPABASE_URL}/rest/v1/{table}", headers=SB_HEADERS, params=params
    )
    resp.raise_for_status()
    return resp.json()


async def sb_patch(table: str, row_id: str, data: dict) -> None:
    resp = await http_client.patch(
        f"{SUPABASE_URL}/rest/v1/{table}",
        headers={**SB_HEADERS, "Prefer": "return=minimal"},
        params={"id": f"eq.{row_id}"},
        json=data,
    )
    resp.raise_for_status()


async def sb_upsert(table: str, row_id: str, data: dict) -> None:
    """PATCH bilan bir xil, lekin qator hali mavjud bo'lmasa ham ishlaydi
    (mavjud bo'lsa yangilaydi, bo'lmasa yaratadi) — biznes_data uchun."""
    resp = await http_client.post(
        f"{SUPABASE_URL}/rest/v1/{table}?on_conflict=id",
        headers={**SB_HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"},
        json={"id": row_id, **data},
    )
    resp.raise_for_status()


async def sb_mutate_and_save(row_id: str, mutator_fn) -> bool:
    """Xavfsiz (race-condition'dan himoyalangan) saqlash: biznes_data
    jadvalidan eng so'nggi ma'lumotni qayta o'qib, shu ustida
    (mutator_fn orqali) o'zgartirish kiritib, keyin saqlaydi — shunda
    shu payt mini app'да xodim kiritgan o'zgarish yo'qolib ketmaydi."""
    try:
        rows = await sb_get("biznes_data", params={"id": f"eq.{row_id}"})
        fresh_data = rows[0]["data"] if rows and isinstance(rows[0].get("data"), dict) else {}
        mutator_fn(fresh_data)
        await sb_upsert("biznes_data", row_id, {"data": fresh_data})
        return True
    except Exception as e:
        log.error("sb_mutate_and_save xatosi (%s): %s", row_id, e)
        return False


async def claim_row(table: str, row_id: str) -> bool:
    """Faqat hozir ham 'pending' bo'lgan qatorni 'processing'ga o'tkazadi.
    Agar boshqa bot nusxasi allaqachon shu qatorni band qilgan bo'lsa,
    bu funksiya False qaytaradi va biz uni qayta yubormaymiz.
    Shu orqali bir nechta bot nusxasi bir vaqtda ishlasa ham,
    xabar hech qachon IKKI MARTA yuborilmaydi."""
    resp = await http_client.patch(
        f"{SUPABASE_URL}/rest/v1/{table}",
        headers={**SB_HEADERS, "Prefer": "return=representation"},
        params={"id": f"eq.{row_id}", "status": "eq.pending"},
        json={"status": "processing"},
    )
    resp.raise_for_status()
    claimed_rows = resp.json()
    return len(claimed_rows) > 0


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ============================================================
# 1) GURUHGA XABARLAR — bot_group_messages
# ============================================================
async def process_group_messages(app: Application) -> None:
    try:
        rows = await sb_get(
            "bot_group_messages",
            params={
                "status": "eq.pending",
                "or": f"(send_at.is.null,send_at.lte.{now_iso()})",
                "order": "created_at.asc",
                "limit": str(BATCH_LIMIT),
            },
        )
    except Exception as e:
        log.error("bot_group_messages fetch xato: %s", e)
        return

    for row in rows:
        row_id = row.get("id")
        try:
            if not await claim_row("bot_group_messages", row_id):
                continue  # boshqa bot nusxasi allaqachon oldi
            text = row.get("text") or ""
            if not text.strip():
                await sb_patch("bot_group_messages", row_id, {"status": "failed"})
                continue
            target_chat_id = row.get("chat_id") or GROUP_CHAT_ID
            await app.bot.send_message(chat_id=target_chat_id, text=text)
            await sb_patch("bot_group_messages", row_id, {"status": "sent"})
            log.info("Guruhga yuborildi: %s -> chat_id=%s", row_id, target_chat_id)
        except Exception as e:
            log.error("Guruhga yuborishda xato (%s): %s", row_id, e)
            try:
                await sb_patch("bot_group_messages", row_id, {"status": "failed"})
            except Exception:
                pass


# ============================================================
# 2) SHAXSIY ESLATMALAR — bot_reminders
# (sekretar uchrashuvlari, kaiten budilnik)
# ============================================================
async def process_reminders(app: Application) -> None:
    try:
        rows = await sb_get(
            "bot_reminders",
            params={
                "status": "eq.pending",
                "remind_at": f"lte.{now_iso()}",
                "order": "remind_at.asc",
                "limit": str(BATCH_LIMIT),
            },
        )
    except Exception as e:
        log.error("bot_reminders fetch xato: %s", e)
        return

    for row in rows:
        row_id = row.get("id")
        try:
            if not await claim_row("bot_reminders", row_id):
                continue  # boshqa bot nusxasi allaqachon oldi
            chat_id = row.get("chat_id")
            text = row.get("text") or ""
            if not chat_id or not text.strip():
                await sb_patch("bot_reminders", row_id, {"status": "failed"})
                continue
            await app.bot.send_message(chat_id=chat_id, text=text)
            await sb_patch("bot_reminders", row_id, {"status": "sent"})
            log.info("Eslatma yuborildi: %s -> chat_id=%s", row_id, chat_id)
        except Exception as e:
            log.error("Eslatma yuborishda xato (%s): %s", row_id, e)
            try:
                await sb_patch("bot_reminders", row_id, {"status": "failed"})
            except Exception:
                pass


# ============================================================
# 3) HUJJAT/XABAR YUBORISH SO'ROVLARI — file_send_requests
# (Hujjatlar bo'limidagi "📤 Yuborish" tugmasi)
# ============================================================
async def process_file_requests(app: Application) -> None:
    try:
        rows = await sb_get(
            "file_send_requests",
            params={
                "status": "eq.pending",
                "order": "created_at.asc",
                "limit": str(BATCH_LIMIT),
            },
        )
    except Exception as e:
        log.error("file_send_requests fetch xato: %s", e)
        return

    for row in rows:
        row_id = row.get("id")
        try:
            if not await claim_row("file_send_requests", row_id):
                continue  # boshqa bot nusxasi allaqachon oldi
            chat_id = row.get("chat_id")
            kind = row.get("kind")
            if not chat_id:
                await sb_patch("file_send_requests", row_id, {"status": "failed"})
                continue

            if kind == "file" and row.get("file_url"):
                await app.bot.send_document(
                    chat_id=chat_id,
                    document=row["file_url"],
                    filename=row.get("file_name") or None,
                )
            elif row.get("message_text"):
                await app.bot.send_message(chat_id=chat_id, text=row["message_text"])
            else:
                await sb_patch("file_send_requests", row_id, {"status": "failed"})
                continue

            await sb_patch("file_send_requests", row_id, {"status": "sent"})
            log.info("Fayl/xabar so'rovi bajarildi: %s -> chat_id=%s", row_id, chat_id)
        except Exception as e:
            log.error("Fayl/xabar yuborishda xato (%s): %s", row_id, e)
            try:
                await sb_patch("file_send_requests", row_id, {"status": "failed"})
            except Exception:
                pass


# ============================================================
# MUAMMOLI MIJOZLAR — davriy tekshiruv xabari (10:15 dan 23:30 gacha,
# har 15 daqiqada), ikki mustaqil tasdiqlash tugmasi bilan
# ============================================================
def generate_check_times() -> list:
    times = []
    start_minutes = 10 * 60 + 15
    end_minutes = 23 * 60 + 30
    cur = start_minutes
    while cur <= end_minutes:
        h, m = divmod(cur, 60)
        times.append(dt_time(hour=h, minute=m, tzinfo=TASHKENT_TZ))
        cur += 15
    return times


def build_check_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Admin panelda hammasi joyida", callback_data="muammoli_check:admin")],
        [InlineKeyboardButton("Telegramda ham hammasi joyida", callback_data="muammoli_check:tg")],
    ])


async def muammoli_check_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    time_label = datetime.now(TASHKENT_TZ).strftime("%H:%M")
    text = (
        f"🔎 TEKSHIRUV — {time_label}\n"
        "📋 Admin panelga qarang — murojaat qilgan muammoli mijoz yo'qmi?\n"
        "👥 Hamkorlar bilan ochilgan telegram guruhlarga qarang — murojaat qilgan hamkor yo'qmi?"
    )
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        usernames = collect_dept_usernames_active(org_nodes, "support")
        if usernames:
            text += "\n\n" + " ".join(f"@{u}" for u in usernames)
    except Exception as e:
        log.error("Support tag xatosi: %s", e)
    try:
        await app.bot.send_message(chat_id=SUPPORT_GROUP_ID, text=text, reply_markup=build_check_keyboard())
        log.info("Tekshiruv xabari yuborildi (%s)", time_label)
    except Exception as e:
        log.error("Tekshiruv xabarini yuborishda xato: %s", e)


async def muammoli_check_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    parts = (query.data or "").split(":", 1)
    btn_type = parts[1] if len(parts) == 2 else ""
    if btn_type not in ("admin", "tg"):
        return

    current_markup = query.message.reply_markup
    still_present = False
    new_rows = []
    if current_markup:
        for row in current_markup.inline_keyboard:
            new_row = [btn for btn in row if btn.callback_data != query.data]
            if len(new_row) != len(row):
                still_present = True
            if new_row:
                new_rows.append(new_row)

    if not still_present:
        return  # allaqachon boshqa kishi tomonidan tasdiqlangan

    user = query.from_user
    name = user.full_name
    username = user.username or ""
    who = f"{name} (@{username})" if username else name
    time_label = datetime.now(TASHKENT_TZ).strftime("%H:%M")

    if btn_type == "admin":
        line = f"✅ {who} — {time_label} holatida admin panelda javob berilmagan murojaat yo'qligini tasdiqladi."
    else:
        line = f"✅ {who} — {time_label} holatida hamkorlar guruhlarida javob berilmagan murojaat yo'qligini tasdiqladi."

    new_text = (query.message.text or "") + "\n\n" + line
    new_markup = InlineKeyboardMarkup(new_rows) if new_rows else None

    try:
        await query.edit_message_text(text=new_text, reply_markup=new_markup)
    except Exception as e:
        log.error("Tekshiruv xabarini tahrirlashda xato: %s", e)


# ============================================================
# MUAMMOLI MIJOZLAR — kunlik eslatma va muddat nazorati
# ============================================================
def parse_tg_field(node: dict) -> list:
    """Xodim tugunida bir nechta username 'tgs' massivida saqlanadi;
    eski yozuvlarda faqat bitta 'tg' maydoni bo'lishi mumkin."""
    tgs = node.get("tgs")
    if isinstance(tgs, list) and tgs:
        return [str(t).strip().lstrip("@") for t in tgs if str(t).strip().lstrip("@")]
    tg = (node.get("tg") or "").strip()
    return [tg.lstrip("@")] if tg else []


def collect_dept_usernames(org_nodes: list, keyword: str) -> list:
    """Org struktura ичida nomi 'keyword' so'zini o'z ichiga olgan tugunni topib,
    uning barcha farzand tugunlaridan (bo'sh bo'lmagan tg maydoni bilan)
    Telegram username'larini yig'ib qaytaradi."""
    if not isinstance(org_nodes, list):
        return []
    dept_ids = [
        n.get("id") for n in org_nodes
        if keyword.lower() in (n.get("name") or "").lower()
    ]
    if not dept_ids:
        return []
    by_parent = {}
    for n in org_nodes:
        by_parent.setdefault(n.get("parentId"), []).append(n)

    usernames = []
    def walk(node_id):
        for child in by_parent.get(node_id, []):
            usernames.extend(parse_tg_field(child))
            walk(child.get("id"))

    for did in dept_ids:
        walk(did)
    return list(dict.fromkeys(usernames))


def collect_support_usernames(org_nodes: list) -> list:
    return collect_dept_usernames(org_nodes, "support")


def _employee_schedule_for_day(node: dict, kun_key: str):
    """Xodimning berilgan kun uchun ish vaqtini qaytaradi: agar shu kunga
    'istisno' (boshqa vaqt) belgilangan bo'lsa o'shani, aks holda umumiy
    ish jadvalini qaytaradi. Natija: (ishBoshlanish, ishTugash)."""
    for exc in node.get("istisnoKunlar") or []:
        if exc.get("kun") == kun_key and exc.get("boshlanish") and exc.get("tugash"):
            return exc["boshlanish"], exc["tugash"]
    return node.get("ishBoshlanish"), node.get("ishTugash")


def _is_currently_working(node: dict) -> bool:
    """Xodimning ish jadvali (kuni + soati, istisno kunlarni hisobga olib)
    hozirgi paytga to'g'ri kelishini tekshiradi. Jadval kiritilmagan
    bo'lsa — False qaytaradi (tag qilinmasin)."""
    now = datetime.now(TASHKENT_TZ)
    kun_key = ISH_KUN_KEYLARI[now.weekday()]
    kunlar = node.get("ishKunlari")
    if kunlar is not None and kun_key not in kunlar:
        return False
    boshlanish, tugash = _employee_schedule_for_day(node, kun_key)
    if not boshlanish or not tugash:
        return False
    try:
        h1, m1 = map(int, boshlanish.split(":"))
        h2, m2 = map(int, tugash.split(":"))
    except Exception:
        return False
    start_min = h1 * 60 + m1
    end_min = h2 * 60 + m2
    now_min = now.hour * 60 + now.minute
    return start_min <= now_min < end_min


def collect_dept_usernames_active(org_nodes: list, keyword: str) -> list:
    """collect_dept_usernames bilan bir xil, lekin faqat HOZIR ish vaqtida
    bo'lgan xodimlarning username'larini qaytaradi (dam olish kunida yoki
    ish vaqtidan tashqarida bo'lganlar chiqarib tashlanadi)."""
    if not isinstance(org_nodes, list):
        return []
    dept_ids = [
        n.get("id") for n in org_nodes
        if keyword.lower() in (n.get("name") or "").lower()
    ]
    if not dept_ids:
        return []
    by_parent = {}
    for n in org_nodes:
        by_parent.setdefault(n.get("parentId"), []).append(n)

    usernames = []
    def walk(node_id):
        for child in by_parent.get(node_id, []):
            if child.get("cardType") != "bolim" and _is_currently_working(child):
                usernames.extend(parse_tg_field(child))
            walk(child.get("id"))

    for did in dept_ids:
        walk(did)
    return list(dict.fromkeys(usernames))


def collect_all_usernames(org_nodes: list) -> list:
    """Org strukturadagi barcha xodimlardan (bo'sh bo'lmagan tg maydoni bilan)
    Telegram username'larini yig'ib qaytaradi."""
    if not isinstance(org_nodes, list):
        return []
    usernames = []
    for n in org_nodes:
        usernames.extend(parse_tg_field(n))
    return list(dict.fromkeys(usernames))


def norm_fio(s: str) -> str:
    """Ism-familiyani solishtirish uchun soddalashtiradi (apostrof/probel farqlarini olib tashlaydi)."""
    s = str(s or "").lower()
    for ch in ("'", "’", "ʻ", "`"):
        s = s.replace(ch, "")
    return " ".join(s.split())


def build_fio_tg_map(org_nodes: list) -> dict:
    """FIO (normallashtirilgan) -> Telegram username xaritasini quradi."""
    mapping = {}
    if not isinstance(org_nodes, list):
        return mapping
    for n in org_nodes:
        fio = (n.get("fio") or "").strip()
        if not fio:
            continue
        tgs = parse_tg_field(n)
        if tgs:
            mapping[norm_fio(fio)] = tgs[0]
    return mapping


async def daily_muammoli_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = rows[0]["data"] if rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []
        open_items = [
            it for it in items
            if not it.get("archived") and it.get("holati") != "hal_qilindi"
        ]

        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        usernames = collect_dept_usernames_active(org_nodes, "support")

        time_label = datetime.now(TASHKENT_TZ).strftime("%H:%M")
        lines = [f"🔔 KUNLIK ESLATMA — Support bo'limiga — {time_label}"]

        if not open_items:
            lines.append("✅ Muammoli mijozlar yo'q")
        else:
            lines.append(f"⚠️ Hozircha muammoli mijozlarda {len(open_items)} ta ochiq murojaat bor")

        if usernames:
            tags = " ".join(f"@{u}" for u in usernames)
            lines.append("")
            lines.append(f"👥 Support bo'limi: {tags}")

        await app.bot.send_message(chat_id=SUPPORT_GROUP_ID, text="\n".join(lines))
        log.info("Kunlik muammoli eslatma yuborildi (%s)", time_label)
    except Exception as e:
        log.error("Kunlik muammoli eslatma xatosi: %s", e)


async def check_overdue_muammoli(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        if not rows:
            return
        muammoli_data = rows[0]["data"]
        if not isinstance(muammoli_data, dict):
            return
        items = muammoli_data.get("items", [])

        org_nodes = []
        try:
            org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
            org_nodes = org_rows[0]["data"] if org_rows else []
        except Exception as e:
            log.error("Org struktura tekshirish xatosi (overdue): %s", e)

        notified_ids = []
        for it in items:
            if it.get("archived") or it.get("holati") == "hal_qilindi":
                continue
            if it.get("overdueNotified"):
                continue
            created_at = it.get("createdAt")
            if not created_at:
                continue
            try:
                created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            except Exception:
                continue
            hours = (datetime.now(timezone.utc) - created_dt).total_seconds() / 3600
            if hours >= MUAMMOLI_MUDDAT_SOAT:
                text = (
                    "⏰ MUDDAT O'TDI (24 soat)!\n"
                    f"🏪 Restoran: {it.get('restoran', '—')}\n"
                    f"🙋 Mijoz: {it.get('ism', '—')}\n"
                    f"📋 Turi: {it.get('turi', 'Boshqa')}\n"
                    "Iltimos, tezroq hal qiling!"
                )
                usernames = collect_dept_usernames_active(org_nodes, "support")
                tags = " ".join(f"@{u}" for u in usernames)
                text += "\n\n@umidpulatov"
                if tags:
                    text += " " + tags
                await app.bot.send_message(chat_id=SUPPORT_GROUP_ID, text=text)
                notified_ids.append(it.get("id"))
        if notified_ids:
            def _mark(fresh_data):
                for fit in fresh_data.get("items", []):
                    if fit.get("id") in notified_ids:
                        fit["overdueNotified"] = True
            ok = await sb_mutate_and_save("muammoli_mijozlar", _mark)
            if ok:
                log.info("Muddati o'tgan murojaat(lar) uchun ogohlantirish yuborildi")
            else:
                log.error("Muddati o'tgan murojaat belgisi saqlanmadi")
    except Exception as e:
        log.error("Muddat tekshirish xatosi: %s", e)


SCREENSHOT_STATE = {"count": 0, "confirmed": False}


def generate_screenshot_times() -> list:
    """10:30 dan 23:30 gacha, har 30 daqiqada."""
    times = []
    start_minutes = 10 * 60 + 30
    end_minutes = 23 * 60 + 30
    cur = start_minutes
    while cur <= end_minutes:
        h, m = divmod(cur, 60)
        times.append(dt_time(hour=h, minute=m, tzinfo=TASHKENT_TZ))
        cur += 30
    return times


async def screenshot_request_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    try:
        SCREENSHOT_STATE["count"] = 0
        SCREENSHOT_STATE["confirmed"] = False
        text = "📸 Iltimos, skrinshot yuboring"
        try:
            org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
            org_nodes = org_rows[0]["data"] if org_rows else []
            usernames = collect_dept_usernames_active(org_nodes, "support")
            if usernames:
                text += "\n\n" + " ".join(f"@{u}" for u in usernames)
        except Exception as e:
            log.error("Skrinshot so'rovida tag xatosi: %s", e)
        await app.bot.send_message(chat_id=SUPPORT_GROUP_ID, text=text)
        log.info("Skrinshot so'rovi yuborildi")
        context.job_queue.run_once(screenshot_followup_check, when=15 * 60)
    except Exception as e:
        log.error("Skrinshot so'rovi xatosi: %s", e)


async def screenshot_followup_check(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    try:
        if not SCREENSHOT_STATE["confirmed"]:
            await app.bot.send_message(
                chat_id=SUPPORT_GROUP_ID,
                text="⚠️ Support bo'limi xodimlari rasm yubormadi\n@umidpulatov",
            )
            log.info("Skrinshot kelmagani haqida ogohlantirish yuborildi")
    except Exception as e:
        log.error("Skrinshot follow-up xatosi: %s", e)


async def handle_support_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.effective_chat or update.effective_chat.id != SUPPORT_GROUP_ID:
        return
    user = update.effective_user
    if not user or not user.username:
        return
    if SCREENSHOT_STATE["confirmed"]:
        return
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        usernames = collect_dept_usernames_active(org_nodes, "support")
    except Exception as e:
        log.error("Org struktura tekshirish xatosi: %s", e)
        return
    if user.username.lstrip("@") not in usernames:
        return

    SCREENSHOT_STATE["count"] += 1
    if SCREENSHOT_STATE["count"] >= 2:
        SCREENSHOT_STATE["confirmed"] = True
        try:
            await update.message.reply_text("✅ Rasmlar qabul qilindi")
        except Exception as e:
            log.error("Rasm tasdiqlash xatosi: %s", e)


async def check_serving_time_reminders(context: ContextTypes.DEFAULT_TYPE) -> None:
    """415 baza'dagi restoranlarning 'ovqat berish vaqti'gacha 1 soat qolganda
    Partnership guruhiga bir martalik eslatma yuboradi (kunlik, restoran uchun)."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        if not rows:
            return
        baza_data = rows[0]["data"]
        if not isinstance(baza_data, dict):
            return
        restoranlar = baza_data.get("restoranlar", [])
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")
        matched = []
        matched_ids = []
        for r in restoranlar:
            if r.get("yashilHamkor"):
                continue
            dan = r.get("berishVaqtiDan")
            if not dan:
                continue
            try:
                h, m = map(int, dan.split(":"))
            except Exception:
                continue
            serving_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            diff_minutes = (serving_dt - now).total_seconds() / 60
            if 89 <= diff_minutes <= 90 and r.get("oxirgiEslatmaSanasi") != today_str:
                matched.append(r)
                matched_ids.append(r.get("id"))
        if matched:
            lines = [
                f"⏰ Berish vaqti boshlanishiga 1 soat 30 daqiqa qolgan {len(matched)} ta restoran bor",
                "Ularni Partnership bo'limidan kirib ko'ring",
                "",
            ]
            for r in matched:
                gacha = r.get("berishVaqtiGacha") or ""
                vaqt = r.get("berishVaqtiDan", "") + (("–" + gacha) if gacha else "")
                lines.append(f"🏪 {r.get('nom','—')} — {vaqt}")
            try:
                org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
                org_nodes = org_rows[0]["data"] if org_rows else []
                usernames = collect_dept_usernames_active(org_nodes, "partnership")
                if usernames:
                    lines.append("")
                    lines.append(" ".join(f"@{u}" for u in usernames))
            except Exception as e:
                log.error("Org struktura tag xatosi: %s", e)
            await app.bot.send_message(chat_id=PARTNERSHIP_GROUP_ID, text="\n".join(lines))
            log.info("Berish vaqti eslatmasi yuborildi: %d ta restoran", len(matched))
        if matched_ids:
            def _mark(fresh_data):
                for fr in fresh_data.get("restoranlar", []):
                    if fr.get("id") in matched_ids:
                        fr["oxirgiEslatmaSanasi"] = today_str
            await sb_mutate_and_save("baza415", _mark)
    except Exception as e:
        log.error("Berish vaqti eslatmasi xatosi: %s", e)


async def check_calling_status_before_serving(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Ovqat berish vaqtidan 2 daqiqa oldin, o'sha vaqtga to'g'ri kelgan
    restoranlar orasida kim tel qilingan, kim qilinmaganini Partnership
    guruhiga yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        if not rows:
            return
        baza_data = rows[0]["data"]
        if not isinstance(baza_data, dict):
            return
        restoranlar = baza_data.get("restoranlar", [])
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")

        # Vaqti 1-2 daqiqa qolgan (bugun hali yuborilmagan) guruhlarni topamiz
        due_times = set()
        for r in restoranlar:
            if r.get("yashilHamkor"):
                continue
            dan = r.get("berishVaqtiDan")
            if not dan:
                continue
            try:
                h, m = map(int, dan.split(":"))
            except Exception:
                continue
            serving_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            diff_minutes = (serving_dt - now).total_seconds() / 60
            if 1 <= diff_minutes <= 2 and r.get("holatXabarSanasi_" + dan) != today_str:
                due_times.add(dan)

        dan_to_ids: dict = {}
        for dan in due_times:
            group = [r for r in restoranlar if r.get("berishVaqtiDan") == dan and not r.get("yashilHamkor")]
            called = [r for r in group if (r.get("qongiroq") or {}).get("lastCalledDate") == today_str]
            not_called = [r for r in group if (r.get("qongiroq") or {}).get("lastCalledDate") != today_str]

            lines = [f"📋 {dan} da berish vaqti boshlangan restoranlarni qaysi biriga telefon qilingan va qilinmaganlari ro'yxati", ""]
            lines.append(f"✅ Tel qilingan: {len(called)} ta")
            for r in called:
                lines.append(f"🏪 {r.get('nom','—')}")
            lines.append("")
            lines.append(f"⏳ Tel qilinmagan: {len(not_called)} ta")
            for r in not_called:
                lines.append(f"🏪 {r.get('nom','—')}")

            try:
                org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
                org_nodes = org_rows[0]["data"] if org_rows else []
                usernames = collect_dept_usernames_active(org_nodes, "partnership")
                if usernames:
                    lines.append("")
                    lines.append(" ".join(f"@{u}" for u in usernames))
            except Exception as e:
                log.error("Org struktura tag xatosi (zvonok1 holati): %s", e)

            await app.bot.send_message(chat_id=PARTNERSHIP_GROUP_ID, text="\n".join(lines))
            log.info("Qo'ng'iroq holati yuborildi: %s", dan)

            dan_to_ids[dan] = [r.get("id") for r in group]

        if dan_to_ids:
            def _mark(fresh_data):
                for fr in fresh_data.get("restoranlar", []):
                    for dan, ids in dan_to_ids.items():
                        if fr.get("id") in ids:
                            fr["holatXabarSanasi_" + dan] = today_str
            await sb_mutate_and_save("baza415", _mark)
    except Exception as e:
        log.error("Qo'ng'iroq holati xatosi: %s", e)


async def daily_calling_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 23:00 (Toshkent) da, 415 baza'dagi barcha restoranlar
    orasida bugun kim tel qilingan, kim qilinmaganini Partnership
    guruhiga umumiy hisobot sifatida yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        if not rows:
            return
        baza_data = rows[0]["data"]
        if not isinstance(baza_data, dict):
            return
        restoranlar = [r for r in baza_data.get("restoranlar", []) if r.get("berishVaqtiDan") and not r.get("yashilHamkor")]
        today = datetime.now(TASHKENT_TZ)
        today_str = today.strftime("%Y-%m-%d")
        today_display = today.strftime("%d.%m.%Y")

        called = [r for r in restoranlar if (r.get("qongiroq") or {}).get("lastCalledDate") == today_str]
        not_called = [r for r in restoranlar if (r.get("qongiroq") or {}).get("lastCalledDate") != today_str]

        lines = [f"📊 KUNLIK QO'NG'IROQLAR HISOBOTI — {today_display}", ""]
        lines.append(f"✅ Tel qilingan: {len(called)} ta")
        for r in called:
            lines.append(f"🏪 {r.get('nom','—')}")
        lines.append("")
        lines.append(f"❌ Tel qilinmagan: {len(not_called)} ta")
        for r in not_called:
            lines.append(f"🏪 {r.get('nom','—')}")

        await app.bot.send_message(chat_id=PARTNERSHIP_GROUP_ID, text="\n".join(lines))
        log.info("Kunlik qo'ng'iroqlar hisoboti yuborildi: %d/%d", len(called), len(restoranlar))
    except Exception as e:
        log.error("Kunlik qo'ng'iroqlar hisoboti xatosi: %s", e)


async def check_zvonok2_call_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Ovqat berish vaqti boshlanganidan 30 daqiqa o'tgach, o'sha vaqtga
    to'g'ri kelgan restoranlarga qo'ng'iroq qilish kerakligini Zvonok 2
    guruhiga yuboradi (Partnership xodimlarini tag qilib)."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        if not rows:
            return
        baza_data = rows[0]["data"]
        if not isinstance(baza_data, dict):
            return
        restoranlar = baza_data.get("restoranlar", [])
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")

        due_times = set()
        for r in restoranlar:
            if r.get("yashilHamkor"):
                continue
            dan = r.get("berishVaqtiDan")
            if not dan:
                continue
            try:
                h, m = map(int, dan.split(":"))
            except Exception:
                continue
            serving_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            diff_minutes = (now - serving_dt).total_seconds() / 60
            if 29 <= diff_minutes <= 30 and r.get("zvonok2XabarSanasi_" + dan) != today_str:
                due_times.add(dan)

        dan_to_ids: dict = {}
        for dan in due_times:
            group = [r for r in restoranlar if r.get("berishVaqtiDan") == dan and not r.get("yashilHamkor")]

            lines = [f"📋 {dan}da berish vaqti boshlangan restoranlar ro'yxatiga qo'ng'iroq qiling", ""]
            for r in group:
                lines.append(f"🏪 {r.get('nom','—')}")

            try:
                org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
                org_nodes = org_rows[0]["data"] if org_rows else []
                usernames = collect_dept_usernames_active(org_nodes, "partnership")
                if usernames:
                    lines.append("")
                    lines.append(" ".join(f"@{u}" for u in usernames))
            except Exception as e:
                log.error("Org struktura tag xatosi (zvonok2): %s", e)

            await app.bot.send_message(chat_id=ZVONOK2_GROUP_ID, text="\n".join(lines))
            log.info("Zvonok 2 qo'ng'iroq eslatmasi yuborildi: %s", dan)

            dan_to_ids[dan] = [r.get("id") for r in group]

        if dan_to_ids:
            def _mark(fresh_data):
                for fr in fresh_data.get("restoranlar", []):
                    for dan, ids in dan_to_ids.items():
                        if fr.get("id") in ids:
                            fr["zvonok2XabarSanasi_" + dan] = today_str
            await sb_mutate_and_save("baza415", _mark)
    except Exception as e:
        log.error("Zvonok 2 qo'ng'iroq eslatmasi xatosi: %s", e)


async def daily_zvonok2_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 23:00 (Toshkent) da, Zvonok 2 bo'yicha bugun kim tel
    qilingan, kim qilinmaganini Zvonok 2 guruhiga umumiy hisobot
    sifatida yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        if not rows:
            return
        baza_data = rows[0]["data"]
        if not isinstance(baza_data, dict):
            return
        restoranlar = [r for r in baza_data.get("restoranlar", []) if r.get("berishVaqtiDan") and not r.get("yashilHamkor")]
        today = datetime.now(TASHKENT_TZ)
        today_str = today.strftime("%Y-%m-%d")
        today_display = today.strftime("%d.%m.%Y")

        called = [r for r in restoranlar if (r.get("qongiroq2") or {}).get("lastCalledDate") == today_str]
        not_called = [r for r in restoranlar if (r.get("qongiroq2") or {}).get("lastCalledDate") != today_str]

        lines = [f"📊 KUNLIK QO'NG'IROQLAR HISOBOTI (Zvonok 2) — {today_display}", ""]
        lines.append(f"✅ Tel qilingan: {len(called)} ta")
        for r in called:
            lines.append(f"🏪 {r.get('nom','—')}")
        lines.append("")
        lines.append(f"❌ Tel qilinmagan: {len(not_called)} ta")
        for r in not_called:
            lines.append(f"🏪 {r.get('nom','—')}")

        await app.bot.send_message(chat_id=ZVONOK2_GROUP_ID, text="\n".join(lines))
        log.info("Zvonok 2 kunlik hisoboti yuborildi: %d/%d", len(called), len(restoranlar))
    except Exception as e:
        log.error("Zvonok 2 kunlik hisoboti xatosi: %s", e)


async def daily_muammoli_hamkorlar_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 23:00 (Toshkent) da, hali 'Restoran bilan gaplashdim'
    bosilmagan (ya'ni faol) Muammoli hamkorlar ro'yxatini Sirly Staff
    guruhiga @kh_nosirov'ni tag qilib yuboradi."""
    app = context.application
    try:
        muammoli_rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = muammoli_rows[0]["data"] if muammoli_rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []

        hamkor_rows = await sb_get("biznes_data", params={"id": "eq.muammoli_hamkorlar"})
        hamkor_data = hamkor_rows[0]["data"] if hamkor_rows and isinstance(hamkor_rows[0].get("data"), dict) else {}
        resetlar = hamkor_data.get("resetlar", {})

        now = datetime.now(timezone.utc)
        d31_cutoff = now - timedelta(days=31)

        restoranlar_set = sorted(set(it.get("restoran") for it in items if it.get("restoran")))

        qualifying = []
        for nom in restoranlar_set:
            reset_iso = resetlar.get(nom)
            reset_dt = None
            if reset_iso:
                try:
                    reset_dt = datetime.fromisoformat(reset_iso.replace("Z", "+00:00"))
                except Exception:
                    reset_dt = None
            cutoff = max(d31_cutoff, reset_dt) if reset_dt else d31_cutoff

            counts: dict = {}
            for it in items:
                if it.get("restoran") != nom or not it.get("turi") or not it.get("createdAt"):
                    continue
                try:
                    created = datetime.fromisoformat(it["createdAt"].replace("Z", "+00:00"))
                except Exception:
                    continue
                if created < cutoff:
                    continue
                counts[it["turi"]] = counts.get(it["turi"], 0) + 1

            max_turi, max_count = None, 0
            for turi, c in counts.items():
                if c > max_count:
                    max_count, max_turi = c, turi
            if max_count >= 3:
                qualifying.append((nom, max_turi, max_count))

        if not qualifying:
            return

        lines = ["⚠️ MUAMMOLI HAMKORLAR — hali gaplashilmagan", ""]
        for nom, turi, count in qualifying:
            lines.append(f"🏪 {nom} — {turi} ({count} marta)")
        lines.append("")
        lines.append("@kh_nosirov")

        await app.bot.send_message(chat_id=MUAMMOLI_HAMKORLAR_REPORT_GROUP_ID, text="\n".join(lines))
        log.info("Muammoli hamkorlar eslatmasi yuborildi: %d ta restoran", len(qualifying))
    except Exception as e:
        log.error("Muammoli hamkorlar eslatmasi xatosi: %s", e)


def _fmt_date_only(iso: str) -> str:
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00")) if "T" in iso else datetime.fromisoformat(iso)
    except Exception:
        return iso
    return dt.astimezone(TASHKENT_TZ).strftime("%d.%m.%Y")


def _bux_signoff(ws, start_row: int, last_col: int) -> None:
    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=10)
    rows = [
        ("Operatsion direktor:", "Umid Pulatov"),
        ("Buxgalter:", "Zilola Maxamatjonova"),
        ("Support bo'limi boshlig'i:", ""),
    ]
    for i, (label, name) in enumerate(rows):
        r = start_row + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font = bold
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        val_cell = ws.cell(row=r, column=4, value=name if name else "_______________________")
        val_cell.font = normal
        val_cell.alignment = Alignment(horizontal="left", vertical="center")

        sign_col = max(7, last_col - 1)
        ws.cell(row=r, column=sign_col, value="Imzo:").font = bold
        ws.merge_cells(start_row=r, start_column=sign_col + 1, end_row=r, end_column=last_col)
        ws.cell(row=r, column=sign_col + 1, value="_______________")
        ws.row_dimensions[r].height = 20


def build_promokod_excel(items: list, today_display: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promokod"

    normal = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E7A4A")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Sana: {today_display}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"] = "PROMOKOD BERISH KERAK — Kunlik hisobot"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color="1E7A4A")
    ws.row_dimensions[2].height = 20

    headers = ["№", "Murojaat sanasi", "Mijoz", "Mijoz raqami", "Restoran", "Sabab", "Support bo'limi xodimi", "Karta raqami"]
    header_row = 4
    ws.row_dimensions[header_row].height = 34
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    for idx, it in enumerate(items, start=1):
        r = header_row + idx
        ws.row_dimensions[r].height = 45
        row_vals = [
            idx,
            _fmt_date_only(it.get("createdAt", "")),
            it.get("ism") or "—",
            it.get("tel") or "—",
            it.get("restoran") or "—",
            it.get("turi") or "—",
            it.get("createdByName") or "—",
            it.get("promokodKartaRaqami") or "—",
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = normal
            cell.border = border
            cell.alignment = center if c_idx in (1, 2, 3) else left

    sign_start = header_row + len(items) + 3
    _bux_signoff(ws, sign_start, 8)

    widths = [5, 14, 14, 15, 16, 36, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pul_berish_excel(items: list, today_display: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pul berish"

    normal = Font(name="Arial", size=10)
    bold = Font(name="Arial", bold=True, size=11)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="E34948")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Sana: {today_display}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:J2")
    ws["A2"] = "RESTORANGA PUL TASHLAB BERISH KERAK — Haftalik hisobot"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color="B0202A")
    ws.row_dimensions[2].height = 20

    headers = ["№", "Murojaat sanasi", "Mijoz", "Mijoz raqami", "Restoran", "Sabab", "Support bo'limi xodimi", "Box soni", "Box narxi", "Jami summa"]
    header_row = 4
    ws.row_dimensions[header_row].height = 34
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    for idx, it in enumerate(items, start=1):
        r = header_row + idx
        ws.row_dimensions[r].height = 45
        soni = it.get("boxSoni") or 1
        narx = it.get("boxSummasi") or 20000
        row_vals = [
            idx,
            _fmt_date_only(it.get("createdAt", "")),
            it.get("ism") or "—",
            it.get("tel") or "—",
            it.get("restoran") or "—",
            it.get("turi") or "—",
            it.get("createdByName") or "—",
            soni,
            narx,
            soni * narx,
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = normal
            cell.border = border
            cell.alignment = center if c_idx in (1, 2, 3, 8, 9, 10) else left
            if c_idx in (9, 10):
                cell.number_format = '#,##0 "so\'m"'

    total_row = header_row + len(items) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
    tot_lbl = ws.cell(row=total_row, column=1, value="JAMI TO'LANADIGAN SUMMA:")
    tot_lbl.font = bold
    tot_lbl.alignment = Alignment(horizontal="right", vertical="center")
    jami_summa = sum((it.get("boxSoni") or 1) * (it.get("boxSummasi") or 20000) for it in items)
    tot_val = ws.cell(row=total_row, column=10, value=jami_summa)
    tot_val.font = Font(name="Arial", bold=True, size=11, color="B0202A")
    tot_val.number_format = '#,##0 "so\'m"'
    tot_val.border = border

    sign_start = total_row + 3
    _bux_signoff(ws, sign_start, 10)

    widths = [5, 14, 13, 15, 16, 36, 20, 10, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_murojaatlar_hisobot_excel(items: list, davr_boshlanish: str, davr_tugash: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Murojaatlar"

    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E7A4A")
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Davr: {davr_boshlanish} — {davr_tugash}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:I2")
    ws["A2"] = "MUROJAATLAR — KUNMA-KUN HISOBOT"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color="1E7A4A")
    ws.row_dimensions[2].height = 20

    headers = ["Sana", "Mijoz", "Restoran", "Muammo turi", "Izoh (tafsilot)", "Qabul qildi", "Holati", "Yechim", "Kim hal qildi"]
    header_row = 4
    ws.row_dimensions[header_row].height = 24
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    # Eng yangi sana birinchi bo'lib chiqishi uchun teskari tartiblaymiz
    sorted_items = sorted(items, key=lambda it: it.get("createdAt") or "", reverse=True)

    for r_off, it in enumerate(sorted_items, start=1):
        r = header_row + r_off
        ws.row_dimensions[r].height = 34
        holati = "Hal qilindi" if it.get("holati") == "hal_qilindi" else "Ochiq"
        row_vals = [
            _fmt_date_only(it.get("createdAt", "")),
            it.get("ism") or "—",
            it.get("restoran") or "—",
            it.get("turi") or "—",
            it.get("izoh") or "—",
            it.get("createdByName") or "—",
            holati,
            it.get("yechimIzohi") or "—",
            it.get("halQildiBy") or "—",
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = normal
            cell.border = border
            cell.alignment = center if c_idx in (1, 3, 7) else left

    widths = [12, 20, 20, 24, 34, 18, 12, 30, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def murojaatlar_hisobot_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/murojaatlar_hisobot buyrug'i — oxirgi 30 kunlik murojaatlar
    (mijoz, restoran, muammo, yechim bilan) Excel faylini yuboradi."""
    await update.message.reply_text("⏳ Tayyorlanmoqda, biroz kuting...")
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = rows[0]["data"] if rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []

        now = datetime.now(TASHKENT_TZ)
        chegara = now - timedelta(days=30)
        filtered = []
        for it in items:
            created_at = it.get("createdAt")
            if not created_at:
                continue
            try:
                created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            except Exception:
                continue
            if created_dt >= chegara:
                filtered.append(it)

        davr_boshlanish = chegara.strftime("%d.%m.%Y")
        davr_tugash = now.strftime("%d.%m.%Y")
        buf = build_murojaatlar_hisobot_excel(filtered, davr_boshlanish, davr_tugash)
        date_tag = now.strftime("%Y%m%d")
        await update.message.reply_document(document=buf, filename=f"murojaatlar_hisobot_{date_tag}.xlsx")
        log.info("Murojaatlar hisoboti yuborildi: %d ta", len(filtered))
    except Exception as e:
        log.error("Murojaatlar hisoboti xatosi: %s", e)
        await update.message.reply_text("⚠️ Xatolik yuz berdi.")


def build_baza415_excel(restoranlar: list, today_display: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "415 baza"

    normal = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E7A4A")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Ma'lumotdagi barcha noyob kontakt rollarini (Menejer, Kassir 1, Kassir 2,
    # Call Center, Buxgalter va h.k.) uchraash tartibida yig'ib olamiz —
    # har bir rol o'z ustuniga ega bo'ladi.
    roles_seen = []
    for r in restoranlar:
        for k in (r.get("kontaktlar") or []):
            rol = (k.get("rol") or "Kontakt").strip()
            if rol not in roles_seen:
                roles_seen.append(rol)

    last_col = 3 + len(roles_seen)
    last_col_letter = get_column_letter(max(last_col, 6))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"Sana: {today_display}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "415 BAZA — Restoranlar ro'yxati"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color="1E7A4A")
    ws.row_dimensions[2].height = 20

    headers = ["№", "Restoran nomi", "Berish vaqti"] + roles_seen
    header_row = 4
    ws.row_dimensions[header_row].height = 20
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    for idx, r in enumerate(restoranlar, start=1):
        row_n = header_row + idx
        ws.row_dimensions[row_n].height = 20
        kontaktlar = r.get("kontaktlar") or []
        role_map: dict = {}
        for k in kontaktlar:
            rol = (k.get("rol") or "Kontakt").strip()
            tel = k.get("tel") or "—"
            ism = k.get("ism")
            val = f"{tel} ({ism})" if ism else tel
            if rol in role_map:
                role_map[rol] = role_map[rol] + ", " + val
            else:
                role_map[rol] = val
        dan = r.get("berishVaqtiDan") or ""
        gacha = r.get("berishVaqtiGacha") or ""
        vaqt = f"{dan}–{gacha}" if dan and gacha else (dan or "—")
        row_vals = [idx, r.get("nom") or "—", vaqt] + [role_map.get(rol, "—") for rol in roles_seen]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_n, column=c_idx, value=val)
            cell.font = normal
            cell.border = border
            cell.alignment = center if c_idx in (1, 3) else left

    widths = [5, 22, 14] + [20] * len(roles_seen)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# JIHOZLAR VA MASHINA TIZIMI (fotograf uchun)
# ============================================================
JIHOZLAR_GROUP_ID = -5477097122
JIHOZLAR_STATE_ID = "jihozlar_davomat"

FOTOGRAF_FIO = "G'aniyev Tohir Abdurafiq o'g'li"
NAZORATCHI_FIO = "Zilola Maxamatjonova"
FOTOGRAF_INFO = {
    "fio": FOTOGRAF_FIO,
    "passport": "AE7299807",
    "tugilgan": "28.02.1999",
    "guvohnoma": "AF0838258",
}

KUNLIK_JIHOZLAR_DATA = [
    [1, "Kamera", "Sony A7R4", "ILCE-7RM4A/WW44785"],
    [2, "Ob'ektiv", "Sony FE 24-105mm f4 G OSS", "SEL24105G/Q"],
    [3, "Sinxronizator", "X2TS", "R211-190627"],
    [4, "Fleshka", "SanDisk 128gb", "SDSDXXD-128G-GN4IN"],
    [5, "Batareyka", "NP-FZ100", "NW101032228EU260415"],
    [6, "Lampa", "AD100PRO2", "CAN ICES-003(B) / NMB-003(B)"],
    [7, "Nosadka softboks uchun", "705-S0000-00", "—"],
    [8, "Sumka", "CANON EOS", "—"],
]
HAFTALIK_JIHOZLAR_DATA = [
    [1, "Shtativ", "QIHE/280/118/100", "—"],
    [2, "Shtativ", "Jmary", "MT-75"],
    [3, "Oktoboks 65 sm", "NS65P", "NW10105596EU2500325"],
    [4, "Derzhatel podsvetka", "ST-RF1", "NW101011O3EU260428"],
]

# Kunlik jarayon: 4 ta video, tartib bo'yicha (kim yuborishi va tasdiq matni)
JIHOZLAR_SEQUENCE = [
    (NAZORATCHI_FIO, "✅ Jihozlar topshirildi va hujjatlar imzolandi. Endi mashinaga o'tasiz."),
    (FOTOGRAF_FIO, "✅ Mashina Tohirga topshirildi."),
    (FOTOGRAF_FIO, "✅ Tohir qaytib keldi. Endi jihozlarni qaytarishga o'tasiz."),
    (NAZORATCHI_FIO, "✅ Jihozlar va mashina qaytarib olindi. Kun yakunlandi."),
]


def _jihoz_doc_header(ws, title, fill_color, xodim_label, xodim_info, sana_dt):
    bold = Font(name="Arial", bold=True, size=11)
    kun_nomlari = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
    kun_nomi = kun_nomlari[sana_dt.weekday()]

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Sana: {sana_dt.strftime('%d.%m.%Y')} ({kun_nomi})"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:D2")
    ws["A2"] = title
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color=fill_color)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:D3")
    ws["A3"] = f"{xodim_label}: {xodim_info['fio']}"
    ws["A3"].font = bold
    ws.row_dimensions[3].height = 18

    ws.merge_cells("A4:D4")
    passport_line = f"Passport: {xodim_info['passport']}   Tug'ilgan sana: {xodim_info['tugilgan']}"
    if xodim_info.get("guvohnoma"):
        passport_line += f"   Haydovchilik guvohnomasi: {xodim_info['guvohnoma']}"
    ws["A4"] = passport_line
    ws["A4"].font = Font(name="Arial", size=10)
    ws.row_dimensions[4].height = 16


def _jihoz_doc_signatures(ws, start_row, fill_color, xodim_info):
    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=10)

    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Topshirilgan vaqt:").font = bold
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="_______________").font = normal
    ws.row_dimensions[r].height = 20

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Qabul qilingan vaqt:").font = bold
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="_______________").font = normal
    ws.row_dimensions[r].height = 20

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="ERTALAB — TOPSHIRISH").font = Font(name="Arial", bold=True, size=10, color=fill_color)
    ws.row_dimensions[r].height = 16

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Xodim (oldim):").font = bold
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r, column=3, value=xodim_info["fio"]).font = normal
    ws.cell(row=r, column=4, value="Imzo: _______________").font = bold
    ws.row_dimensions[r].height = 26

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Nazoratchi (berdim):").font = bold
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r, column=3, value=NAZORATCHI_FIO).font = normal
    ws.cell(row=r, column=4, value="Imzo: _______________").font = bold
    ws.row_dimensions[r].height = 26

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="KECHQURUN — QABUL QILISH").font = Font(name="Arial", bold=True, size=10, color=fill_color)
    ws.row_dimensions[r].height = 16

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Xodim (qaytardim):").font = bold
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r, column=3, value=xodim_info["fio"]).font = normal
    ws.cell(row=r, column=4, value="Imzo: _______________").font = bold
    ws.row_dimensions[r].height = 26

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Nazoratchi (qabul qildim):").font = bold
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r, column=3, value=NAZORATCHI_FIO).font = normal
    ws.cell(row=r, column=4, value="Imzo: _______________").font = bold
    ws.row_dimensions[r].height = 26


def build_jihozlar_excel(sana_dt, title, fill_color, data) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dalolatnoma"

    normal = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _jihoz_doc_header(ws, title, fill_color, "Uskunalardan foydalanuvchi xodim", FOTOGRAF_INFO, sana_dt)

    header_row = 6
    headers = ["№", "Nomi", "Modeli", "Seriya raqami", "Topshirdi\n(✓)", "Qabul qildi\n(✓)"]
    ws.row_dimensions[header_row].height = 30
    header_fill = PatternFill("solid", fgColor=fill_color)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    for r_off, row in enumerate(data, start=1):
        r = header_row + r_off
        ws.row_dimensions[r].height = 24
        full_row = row + ["", ""]
        for c_idx, val in enumerate(full_row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = normal
            cell.border = border
            cell.alignment = center if c_idx in (1, 5, 6) else left

    sign_start = header_row + len(data) + 3
    _jihoz_doc_signatures(ws, sign_start, fill_color, FOTOGRAF_INFO)

    widths = [5, 26, 26, 28, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_mashina_excel(sana_dt) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dalolatnoma"

    bold = Font(name="Arial", bold=True, size=11)
    fill_color = "2a5ea8"

    _jihoz_doc_header(ws, "MASHINANI TOPSHIRISH VA QABUL QILISH DALOLATNOMASI", fill_color,
                       "Mashinadan foydalanuvchi xodim", FOTOGRAF_INFO, sana_dt)

    ws.merge_cells("A5:D5")
    ws["A5"] = "Chevrolet Spark  |  Davlat raqami: 01H131XC  |  Rangi: Yashil  |  Ishlab chiqarilgan yili: 2015"
    ws["A5"].font = Font(name="Arial", size=10)
    ws.row_dimensions[5].height = 16

    ws.merge_cells("A6:D6")
    ws["A6"] = "Kuzov raqami: XWBMA481JFA512764  |  Dvigatel raqami: B10D1201370KD3"
    ws["A6"].font = Font(name="Arial", size=10)
    ws.row_dimensions[6].height = 16

    r = 8
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Mashina topshirildi:").font = bold
    ws.cell(row=r, column=3, value="☐").font = Font(name="Arial", size=18)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 24

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Mashina qabul qilindi:").font = bold
    ws.cell(row=r, column=3, value="☐").font = Font(name="Arial", size=18)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 24

    _jihoz_doc_signatures(ws, r + 2, fill_color, FOTOGRAF_INFO)

    widths = [22, 20, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def daily_jihozlar_hujjat(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 23:00 (Toshkent) da, agar ERTAGA yakshanba bo'lmasa, kunlik
    jihozlar va mashina hujjatlarini JIHOZLAR guruhiga yuboradi."""
    app = context.application
    try:
        now = datetime.now(TASHKENT_TZ)
        tomorrow = now + timedelta(days=1)
        if tomorrow.weekday() == 6:  # Yakshanba
            return

        kunlik_buf = build_jihozlar_excel(tomorrow, "KUNLIK JIHOZLARNI TOPSHIRISH VA QABUL QILISH DALOLATNOMASI", "1E7A4A", KUNLIK_JIHOZLAR_DATA)
        mashina_buf = build_mashina_excel(tomorrow)
        date_tag = tomorrow.strftime("%Y%m%d")

        text = (
            f"📋 Ertangi kun ({tomorrow.strftime('%d.%m.%Y')}) uchun jihozlar va mashina hujjatlari.\n\n"
            "Zilola, ertalab jihozlarni va tasdiqlangan hujjatlarni ikki tomonlama imzolab, "
            "video xabar yuboring."
        )
        await app.bot.send_message(chat_id=JIHOZLAR_GROUP_ID, text=text)
        await app.bot.send_document(chat_id=JIHOZLAR_GROUP_ID, document=kunlik_buf, filename=f"kunlik_jihozlar_{date_tag}.xlsx")
        await app.bot.send_document(chat_id=JIHOZLAR_GROUP_ID, document=mashina_buf, filename=f"mashina_{date_tag}.xlsx")

        def _reset(data):
            data[tomorrow.strftime("%Y-%m-%d")] = {"step": 0}
        await sb_mutate_and_save(JIHOZLAR_STATE_ID, _reset)

        log.info("Kunlik jihozlar hujjati yuborildi: %s", tomorrow.strftime("%d.%m.%Y"))
    except Exception as e:
        log.error("Kunlik jihozlar hujjati xatosi: %s", e)


async def weekly_jihozlar_hujjat(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har shanba, haftalik jihozlar hujjatini JIHOZLAR guruhiga yuboradi."""
    app = context.application
    try:
        now = datetime.now(TASHKENT_TZ)
        buf = build_jihozlar_excel(now, "HAFTALIK JIHOZLARNI TOPSHIRISH VA QABUL QILISH DALOLATNOMASI", "8a5c00", HAFTALIK_JIHOZLAR_DATA)
        date_tag = now.strftime("%Y%m%d")
        text = (
            "📋 Haftalik jihozlar hujjati (mashina bagajidagi jihozlar).\n\n"
            "Tohir, ko'chaga chiqib, mashina yonida bagajdagi jihozlarni video oling va guruhga yuboring."
        )
        await app.bot.send_message(chat_id=JIHOZLAR_GROUP_ID, text=text)
        await app.bot.send_document(chat_id=JIHOZLAR_GROUP_ID, document=buf, filename=f"haftalik_jihozlar_{date_tag}.xlsx")
        log.info("Haftalik jihozlar hujjati yuborildi")
    except Exception as e:
        log.error("Haftalik jihozlar hujjati xatosi: %s", e)


async def test_jihozlar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("⏳ Tekshirilmoqda, biroz kuting...")
    await daily_jihozlar_hujjat(context)
    await weekly_jihozlar_hujjat(context)
    await update.message.reply_text("✅ Tayyor! Jihozlar guruhini tekshiring — 3 ta hujjat o'sha yerga yuborildi.")


async def handle_jihozlar_video(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Jihozlar guruhiga kelgan video xabarlarni (krujok) kuzatib, kunlik
    4 bosqichli jarayonni (Zilola->Tohir->Tohir->Zilola) tekshiradi."""
    if not update.effective_chat or update.effective_chat.id != JIHOZLAR_GROUP_ID:
        return
    user = update.effective_user
    if not user or not user.username:
        return
    try:
        username = user.username.lstrip("@")
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []

        sender_fio = None
        for n in org_nodes:
            if username in parse_tg_field(n):
                sender_fio = (n.get("fio") or "").strip()
                break
        if sender_fio not in (FOTOGRAF_FIO, NAZORATCHI_FIO):
            return

        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")

        state_rows = await sb_get("biznes_data", params={"id": f"eq.{JIHOZLAR_STATE_ID}"})
        state = state_rows[0]["data"] if state_rows and isinstance(state_rows[0].get("data"), dict) else {}
        day_state = state.get(today_str) or {"step": 0}
        step = day_state.get("step", 0)

        if step >= 4:
            return

        expected_fio, confirm_text = JIHOZLAR_SEQUENCE[step]
        if sender_fio != expected_fio:
            return

        day_state["step"] = step + 1
        day_state[f"video{step+1}_at"] = now.isoformat()
        day_state[f"video{step+1}_by"] = username

        def _mark(data):
            data[today_str] = day_state
        ok = await sb_mutate_and_save(JIHOZLAR_STATE_ID, _mark)
        if not ok:
            log.error("Jihozlar davomat saqlanmadi")
            return

        await context.application.bot.send_message(chat_id=JIHOZLAR_GROUP_ID, text=confirm_text)
        log.info("Jihozlar video qadam %d tasdiqlandi: %s", step + 1, username)
    except Exception as e:
        log.error("Jihozlar video ishlov berish xatosi: %s", e)


async def check_jihozlar_deadline(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Soat 11:00 va 20:00 da, kunlik jihozlar jarayoni belgilangan
    bosqichda qolib ketgan bo'lsa, @umidpulatov'ni tag qilib ogohlantiradi."""
    app = context.application
    try:
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")
        hour = now.hour

        state_rows = await sb_get("biznes_data", params={"id": f"eq.{JIHOZLAR_STATE_ID}"})
        state = state_rows[0]["data"] if state_rows and isinstance(state_rows[0].get("data"), dict) else {}
        day_state = state.get(today_str)
        if not day_state:
            return
        step = day_state.get("step", 0)

        if hour == 11 and step < 2 and not day_state.get("alert11_sent"):
            labels = ["Jihozlar topshirilishi (Zilola)", "Mashina topshirilishi (Tohir)"]
            missing = labels[step:2]
            text = "⚠️ Soat 11:00 bo'ldi, lekin quyidagilar hali tasdiqlanmagan:\n" + "\n".join(f"• {m}" for m in missing) + "\n\n@umidpulatov"
            await app.bot.send_message(chat_id=JIHOZLAR_GROUP_ID, text=text)
            day_state["alert11_sent"] = True
            def _mark11(data):
                data[today_str] = day_state
            await sb_mutate_and_save(JIHOZLAR_STATE_ID, _mark11)

        if hour == 20 and step < 4 and not day_state.get("alert20_sent"):
            labels = ["Jihozlar topshirilishi (Zilola)", "Mashina topshirilishi (Tohir)", "Tohirning qaytishi", "Jihozlar/mashina qaytarib olinishi (Zilola)"]
            missing = labels[step:4]
            text = "⚠️ Soat 20:00 bo'ldi, lekin quyidagilar hali tasdiqlanmagan:\n" + "\n".join(f"• {m}" for m in missing) + "\n\n@umidpulatov"
            await app.bot.send_message(chat_id=JIHOZLAR_GROUP_ID, text=text)
            day_state["alert20_sent"] = True
            def _mark20(data):
                data[today_str] = day_state
            await sb_mutate_and_save(JIHOZLAR_STATE_ID, _mark20)
    except Exception as e:
        log.error("Jihozlar muddat tekshiruvi xatosi: %s", e)


async def daily_baza415_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 10:00 (Toshkent) da, 415 baza'dagi barcha restoranlar
    ro'yxatini Excel fayl sifatida tegishli guruhga yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        baza_data = rows[0]["data"] if rows else {}
        restoranlar = baza_data.get("restoranlar", []) if isinstance(baza_data, dict) else []
        restoranlar = sorted(restoranlar, key=lambda r: (r.get("nom") or "").lower())
        today_display = datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y")
        date_tag = datetime.now(TASHKENT_TZ).strftime("%Y%m%d")

        buf = build_baza415_excel(restoranlar, today_display)
        await app.bot.send_document(
            chat_id=BAZA415_REPORT_GROUP_ID,
            document=buf,
            filename=f"415_baza_{date_tag}.xlsx",
        )
        log.info("415 baza hisoboti yuborildi: %d ta restoran", len(restoranlar))
    except Exception as e:
        log.error("415 baza hisoboti xatosi: %s", e)


async def check_yashil_hamkor_auto(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni tekshiradi: faollashtirilgan sanasidan 15 kun o'tgan, hali
    Yashil hamkor bo'lmagan restoranlarni. Agar shu 15 kun ichida bitta
    muammo turi 3 martaga yetmagan bo'lsa — avtomatik Yashil hamkor qiladi
    va 415 baza guruhiga xabar beradi."""
    app = context.application
    try:
        baza_rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        baza_data = baza_rows[0]["data"] if baza_rows else {}
        restoranlar = baza_data.get("restoranlar", []) if isinstance(baza_data, dict) else []

        muammoli_rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = muammoli_rows[0]["data"] if muammoli_rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []

        now = datetime.now(timezone.utc)
        newly_green_ids = []

        for r in restoranlar:
            if r.get("yashilHamkor"):
                continue
            faol_iso = r.get("faollashtirilganSana")
            if not faol_iso:
                continue
            try:
                faol_dt = datetime.fromisoformat(faol_iso.replace("Z", "+00:00"))
            except Exception:
                continue
            kun_otgan = (now - faol_dt).days
            if kun_otgan < 15:
                continue

            window_end = faol_dt + timedelta(days=15)
            counts: dict = {}
            for it in items:
                if it.get("restoran") != r.get("nom") or not it.get("turi") or not it.get("createdAt"):
                    continue
                try:
                    created = datetime.fromisoformat(it["createdAt"].replace("Z", "+00:00"))
                except Exception:
                    continue
                if not (faol_dt <= created <= window_end):
                    continue
                counts[it["turi"]] = counts.get(it["turi"], 0) + 1

            max_count = max(counts.values()) if counts else 0
            if max_count >= 3:
                continue  # muammoli hamkor bo'lgan, yashil bo'lolmaydi

            newly_green_ids.append(r.get("id"))

            lines = ["🟢 YASHIL HAMKOR STATUSIGA O'TDI", "", f"🏪 {r.get('nom','—')}", "📅 15 kunlik kuzatuv davomida yashil hamkor statusini oldi", ""]
            if counts:
                lines.append("Kuzatuv davrida uchragan muammolar:")
                for turi, c in counts.items():
                    lines.append(f"• {turi} — {c} marta")
            else:
                lines.append("Kuzatuv davrida hech qanday muammo bo'lmadi.")
            lines.append("")
            lines.append(datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y"))

            await app.bot.send_message(chat_id=BAZA415_REPORT_GROUP_ID, text="\n".join(lines))
            log.info("Avtomatik yashil hamkor: %s", r.get("nom"))

        if newly_green_ids:
            yashil_sana = now.isoformat()
            def _mark(fresh_data):
                for fr in fresh_data.get("restoranlar", []):
                    if fr.get("id") in newly_green_ids:
                        fr["yashilHamkor"] = True
                        fr["yashilBelgilaganSana"] = yashil_sana
            await sb_mutate_and_save("baza415", _mark)
    except Exception as e:
        log.error("Yashil hamkor avtomatik tekshiruvi xatosi: %s", e)


async def noaktiv_hamkorlar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/noaktiv_hamkorlar buyrug'i — 'Ovqat berish vaqti' kiritilmagan
    (hali aktiv bo'lmagan) restoranlar ro'yxatini 415 baza guruhiga yuboradi."""
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.baza415"})
        baza_data = rows[0]["data"] if rows else {}
        restoranlar = baza_data.get("restoranlar", []) if isinstance(baza_data, dict) else []
        noaktiv = sorted(
            [r for r in restoranlar if not r.get("berishVaqtiDan")],
            key=lambda r: (r.get("nom") or "").lower(),
        )

        lines = ["⚪️ NOAKTIV HAMKORLAR — \"Ovqat berish vaqti\" kiritilmagan", ""]
        if noaktiv:
            for r in noaktiv:
                lines.append(f"🏪 {r.get('nom','—')}")
            lines.append("")
            lines.append(f"Jami: {len(noaktiv)} ta")
        else:
            lines.append("Barcha restoranlarda berish vaqti kiritilgan. 🎉")

        await context.application.bot.send_message(chat_id=BAZA415_REPORT_GROUP_ID, text="\n".join(lines))
        await update.message.reply_text("✅ Tayyor! 415 baza guruhini tekshiring.")
    except Exception as e:
        log.error("Noaktiv hamkorlar buyrug'i xatosi: %s", e)
        await update.message.reply_text("⚠️ Xatolik yuz berdi.")


async def daily_promokod_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 18:00 (Toshkent) da, hal bo'lmagan murojaatlar orasida
    'Promokod berish kerak' deb belgilanganlarning Excel faylini
    Buxgalteriya guruhiga yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = rows[0]["data"] if rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []
        open_items = [it for it in items if not it.get("archived") and it.get("holati") != "hal_qilindi"]

        promokod_items = [it for it in open_items if it.get("promokodKerak")]

        today_display = datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y")
        date_tag = datetime.now(TASHKENT_TZ).strftime("%Y%m%d")

        text = "📋 Promokod berish kerak bo'lgan mijozlar ro'yxati quyidagi Excel faylda"
        try:
            org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
            org_nodes = org_rows[0]["data"] if org_rows else []
            usernames = collect_dept_usernames_active(org_nodes, "support")
            if usernames:
                text += "\n\n" + " ".join(f"@{u}" for u in usernames)
        except Exception as e:
            log.error("Org struktura tag xatosi (promokod): %s", e)
        await app.bot.send_message(chat_id=BUXGALTERIYA_PROMOKOD_GROUP_ID, text=text)

        promokod_buf = build_promokod_excel(promokod_items, today_display)
        await app.bot.send_document(
            chat_id=BUXGALTERIYA_PROMOKOD_GROUP_ID,
            document=promokod_buf,
            filename=f"promokod_{date_tag}.xlsx",
        )

        log.info("Kunlik promokod hisoboti yuborildi: %d ta", len(promokod_items))
    except Exception as e:
        log.error("Kunlik promokod hisoboti xatosi: %s", e)


async def weekly_pul_berish_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har juma 18:00 (Toshkent) da, hal bo'lmagan murojaatlar orasida
    'Restoranga pul tashlab berish kerak' deb belgilanganlarning Excel
    faylini Buxgalteriya guruhiga yuboradi."""
    app = context.application
    try:
        rows = await sb_get("biznes_data", params={"id": "eq.muammoli_mijozlar"})
        muammoli_data = rows[0]["data"] if rows else {}
        items = muammoli_data.get("items", []) if isinstance(muammoli_data, dict) else []
        open_items = [it for it in items if not it.get("archived") and it.get("holati") != "hal_qilindi"]

        pul_items = [it for it in open_items if it.get("tasdiqlamadiLekinBekor")]

        today_display = datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y")
        date_tag = datetime.now(TASHKENT_TZ).strftime("%Y%m%d")

        text = "📋 Restoran tasdiqlamagan sifatsiz taom uchun to'lov ro'yxati quyidagi Excel faylda"
        await app.bot.send_message(chat_id=BUXGALTERIYA_PROMOKOD_GROUP_ID, text=text)

        pul_buf = build_pul_berish_excel(pul_items, today_display)
        await app.bot.send_document(
            chat_id=BUXGALTERIYA_PROMOKOD_GROUP_ID,
            document=pul_buf,
            filename=f"restoranga_pul_berish_{date_tag}.xlsx",
        )

        log.info("Haftalik pul berish hisoboti yuborildi: %d ta", len(pul_items))
    except Exception as e:
        log.error("Haftalik pul berish hisoboti xatosi: %s", e)


# ============================================================
# ISH DAVOMATI — krujok (video note) orqali tasdiqlash
#
# Race condition oldini olish uchun har bir jarayon O'ZINING
# alohida Supabase qatoriga yozadi (bir-birining ustidan
# yozib yubormasligi uchun):
#   - ATTENDANCE_KELGANLAR_ID   -> handle_staff_video_note yozadi
#   - ATTENDANCE_REMINDED_ID    -> check_ish_boshlanish_reminder yozadi
#   - ATTENDANCE_KELMAGAN_ID    -> check_ish_kelmagan yozadi
# ============================================================
ATTENDANCE_KELGANLAR_ID = "ish_davomat_kelganlar"
ATTENDANCE_REMINDED_ID = "ish_davomat_reminded"
ATTENDANCE_KELMAGAN_ID = "ish_davomat_kelmagan"

KUN_NOMLARI_UZ = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
ISH_KUN_KEYLARI = ["dush", "sesh", "chor", "pay", "juma", "shan", "yak"]


def _org_employees_with_schedule(org_nodes: list) -> list:
    """ishBoshlanish maydoni to'ldirilgan barcha XODIM (bo'lim emas) tugunlarini qaytaradi."""
    if not isinstance(org_nodes, list):
        return []
    return [n for n in org_nodes if n.get("ishBoshlanish") and n.get("cardType") != "bolim"]


def _today_ish_kun_key() -> str:
    return ISH_KUN_KEYLARI[datetime.now(TASHKENT_TZ).weekday()]


def _find_dept_name(node: dict, nodes_by_id: dict) -> str:
    """Xodim tugunidan yuqoriga qarab, birinchi 'bolim' turidagi ota tugunni topadi."""
    parent_id = node.get("parentId")
    seen = set()
    while parent_id and parent_id not in seen:
        seen.add(parent_id)
        parent = nodes_by_id.get(parent_id)
        if not parent:
            break
        if parent.get("cardType") == "bolim":
            return parent.get("name") or "Boshqa bo'lim"
        parent_id = parent.get("parentId")
    return "Boshqa bo'lim"


def _group_employees_by_dept(employees: list, org_nodes: list) -> dict:
    nodes_by_id = {n.get("id"): n for n in org_nodes if n.get("id")}
    by_dept: dict = {}
    for n in employees:
        dept = _find_dept_name(n, nodes_by_id)
        by_dept.setdefault(dept, []).append(n)
    return by_dept


IT_BOLIMI_GROUP_ID = -5264364602


async def daily_it_report_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni (Yakshanbadan tashqari) 22:30 (Toshkent) da, IT bo'limi
    guruhiga, Org strukturadagi IT bo'limi xodimlarini tag qilib,
    kunlik hisobot so'raydi."""
    app = context.application
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        usernames = collect_dept_usernames(org_nodes, "it bo'limi")
        tag_line = " ".join(f"@{u}" for u in usernames) if usernames else ""

        lines = [
            "📋 Kunlik hisobot",
            "",
            tag_line,
            "",
            "Kunlik bag va tasklar ro'yxatini shakillantirganingizni skrinshoti va shu faylga kirish uchun link yuboring.",
            "",
            "Hech qaysi bag va task qolib ketmaganini tasdiqlang.",
        ]
        await app.bot.send_message(chat_id=IT_BOLIMI_GROUP_ID, text="\n".join(lines))
        log.info("IT bo'limi kunlik hisobot so'rovi yuborildi")
    except Exception as e:
        log.error("IT bo'limi kunlik hisobot xatosi: %s", e)


async def daily_tomorrow_schedule_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 22:00 (Toshkent) da, ertaga ishga chiqishi kerak bo'lgan
    xodimlar ro'yxatini (bo'limlar bo'yicha, boshlanish-tugash vaqti bilan)
    HR guruhiga yuboradi. Shu bilan birga, bo'lim talab qilgan soatlarda
    xodimsiz qolgan vaqt (tuynuk) borligini ham tekshirib, xabar oxiriga
    qo'shib, @umidpulatov'ni tag qiladi."""
    app = context.application
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        if not isinstance(org_nodes, list):
            org_nodes = []

        tomorrow = datetime.now(TASHKENT_TZ) + timedelta(days=1)
        tomorrow_kun = ISH_KUN_KEYLARI[tomorrow.weekday()]
        tomorrow_display = tomorrow.strftime("%d.%m.%Y") + f" ({KUN_NOMLARI_UZ[tomorrow.weekday()]})"

        employees = _org_employees_with_schedule(org_nodes)
        active_tomorrow = [n for n in employees if n.get("ishKunlari") is None or tomorrow_kun in n.get("ishKunlari")]

        lines = [f"📅 {tomorrow_display} — Ertangi ish jadvali", ""]
        if active_tomorrow:
            by_dept = _group_employees_by_dept(active_tomorrow, org_nodes)
            for dept, members in sorted(by_dept.items()):
                lines.append(f"🏢 {dept}")
                for n in members:
                    fio = n.get("fio") or n.get("name") or "—"
                    boshlanish, tugash = _employee_schedule_for_day(n, tomorrow_kun)
                    lines.append(f"👤 {fio} — {boshlanish or '—'}–{tugash or '—'}")
                lines.append("")
        else:
            lines.append("Hech kim ishga chiqmaydi.")
            lines.append("")

        # Bo'limlarda xodimsiz qolgan vaqt (tuynuk) borligini tekshiramiz
        depts = [n for n in org_nodes if n.get("cardType") == "bolim" and n.get("ishBoshlanish") and n.get("ishTugash")]
        problem_lines = []
        for dept in depts:
            dept_kunlar = dept.get("ishKunlari")
            if dept_kunlar is not None and tomorrow_kun not in dept_kunlar:
                continue  # bo'lim ertaga umuman ishlamaydi
            try:
                dept_start = _minutes(dept["ishBoshlanish"])
                dept_end = _minutes(dept["ishTugash"])
            except Exception:
                continue

            dept_employees = _find_department_employees(dept.get("id"), org_nodes)
            gaps = _find_coverage_gaps(dept_start, dept_end, dept_employees, tomorrow_kun)
            if gaps:
                dept_name = dept.get("name") or "Bo'lim"
                gap_str = ", ".join(f"{_fmt_minutes(s)}–{_fmt_minutes(e)}" for s, e in gaps)
                problem_lines.append(f"🏢 {dept_name} — xodimsiz qolgan vaqt: {gap_str}")

        if problem_lines:
            lines.append("⚠️ XODIMSIZ QOLADIGAN BO'LIMLAR")
            lines.extend(problem_lines)
            lines.append("")
            lines.append("@umidpulatov")

        if not active_tomorrow and not problem_lines:
            return

        await app.bot.send_message(chat_id=HR_GROUP_ID, text="\n".join(lines))
        log.info("Ertangi ish jadvali yuborildi: %d xodim, %d muammo", len(active_tomorrow), len(problem_lines))
    except Exception as e:
        log.error("Ertangi ish jadvali xatosi: %s", e)


def _minutes(t: str) -> int:
    h, m = map(int, t.split(":"))
    return h * 60 + m


def _fmt_minutes(m: int) -> str:
    return f"{m // 60:02d}:{m % 60:02d}"


def _find_department_employees(dept_id: str, org_nodes: list) -> list:
    """Bo'lim tuguni ostidagi barcha XODIM avlodlarini (rekursiv) topadi."""
    by_parent: dict = {}
    for n in org_nodes:
        by_parent.setdefault(n.get("parentId"), []).append(n)
    result = []
    def walk(node_id):
        for child in by_parent.get(node_id, []):
            if child.get("cardType") != "bolim":
                result.append(child)
            walk(child.get("id"))
    walk(dept_id)
    return result


def _find_coverage_gaps(dept_start: int, dept_end: int, employees: list, kun_key: str) -> list:
    """Bo'lim talab qilgan [dept_start, dept_end) oralig'ida, xodimlar
    jadvali orqali QOPLANMAGAN (tuynuk) segmentlarni topadi."""
    if dept_end <= dept_start:
        return []
    covered = [False] * (dept_end - dept_start)
    for emp in employees:
        kunlar = emp.get("ishKunlari")
        if kunlar is not None and kun_key not in kunlar:
            continue
        boshlanish, tugash = _employee_schedule_for_day(emp, kun_key)
        if not boshlanish or not tugash:
            continue
        try:
            es = _minutes(boshlanish)
            ee = _minutes(tugash)
        except Exception:
            continue
        for i in range(max(es, dept_start), min(ee, dept_end)):
            covered[i - dept_start] = True

    gaps = []
    i = 0
    n = len(covered)
    while i < n:
        if not covered[i]:
            start = i
            while i < n and not covered[i]:
                i += 1
            gaps.append((dept_start + start, dept_start + i))
        else:
            i += 1
    return gaps


async def check_ish_boshlanish_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Kunning eng erta ish boshlanish vaqtiga 10 daqiqa qolganda — HR guruhiga
    BUGUNGI TO'LIQ jadval (barcha bo'lim, barcha xodim, dam olish kuni bilan)
    yuboriladi. Boshqa (keyinroq) vaqtlar uchun esa faqat o'sha vaqtga to'g'ri
    keladigan xodimlar, o'z bo'limi bilan, qisqa xabar sifatida yuboriladi.
    Har xodimning bugungi ISTISNO kuni bo'lsa, o'sha vaqt ishlatiladi."""
    app = context.application
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        employees = _org_employees_with_schedule(org_nodes)
        if not employees:
            return

        today_kun = _today_ish_kun_key()
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")
        today_display = now.strftime("%d.%m.%Y") + f" ({KUN_NOMLARI_UZ[now.weekday()]})"

        state_rows = await sb_get("biznes_data", params={"id": f"eq.{ATTENDANCE_REMINDED_ID}"})
        state = state_rows[0]["data"] if state_rows and isinstance(state_rows[0].get("data"), dict) else {}
        reminded = state.setdefault(today_str, [])

        active_today = [n for n in employees if n.get("ishKunlari") is None or today_kun in n.get("ishKunlari")]
        if not active_today:
            return

        # Har xodimning BUGUNGI (istisno kunlarni hisobga olgan) boshlanish vaqti
        today_start = {}
        for n in active_today:
            b, _ = _employee_schedule_for_day(n, today_kun)
            if b:
                today_start[n.get("id") or id(n)] = b

        times_today = sorted(set(today_start.values()))
        if not times_today:
            return
        earliest = times_today[0]

        due_times = []
        for vaqt in times_today:
            try:
                h, m = map(int, vaqt.split(":"))
            except Exception:
                continue
            start_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            diff_minutes = (start_dt - now).total_seconds() / 60
            if 9 <= diff_minutes <= 10 and vaqt not in reminded:
                due_times.append(vaqt)

        if not due_times:
            return

        instruction = (
            "📹 Ishga kelganingizni video xabar (krujok) orqali tasdiqlang — "
            "video'da ofis kompyuterining sana va soatini, hamda o'zingizni "
            "tasdiqlaydigan biror belgi (masalan qo'l bilan ishora) ko'rsating."
        )

        for vaqt in due_times:
            if vaqt == earliest:
                # Bugungi TO'LIQ jadval, bo'limlar bo'yicha guruhlangan
                by_dept = _group_employees_by_dept(employees, org_nodes)
                lines = [f"📅 {today_display} — Bugungi ish jadvali", ""]
                for dept, members in sorted(by_dept.items()):
                    lines.append(f"🏢 {dept}")
                    for n in members:
                        fio = n.get("fio") or n.get("name") or "—"
                        kunlar = n.get("ishKunlari")
                        if kunlar is not None and today_kun not in kunlar:
                            lines.append(f"👤 {fio} — Dam olish kuni")
                        else:
                            b, t = _employee_schedule_for_day(n, today_kun)
                            lines.append(f"👤 {fio} — {b or '—'}–{t or '—'}")
                    lines.append("")
                lines.append(instruction)
                text = "\n".join(lines)
            else:
                group = [n for n in active_today if today_start.get(n.get("id") or id(n)) == vaqt]
                by_dept = _group_employees_by_dept(group, org_nodes)
                lines = [f"⏰ Ish vaqti boshlanishiga 10 daqiqa qolgan xodimlar ({vaqt}):", ""]
                for dept, members in sorted(by_dept.items()):
                    lines.append(f"🏢 {dept}")
                    for n in members:
                        fio = n.get("fio") or n.get("name") or "—"
                        tgs = parse_tg_field(n)
                        tag = f" (@{tgs[0]})" if tgs else ""
                        lines.append(f"👤 {fio}{tag}")
                    lines.append("")
                lines.append(instruction)
                text = "\n".join(lines)

            await app.bot.send_message(chat_id=HR_GROUP_ID, text=text)
            reminded.append(vaqt)
            log.info("Ish boshlanish eslatmasi yuborildi: %s", vaqt)

        await sb_upsert("biznes_data", ATTENDANCE_REMINDED_ID, {"data": state})
    except Exception as e:
        log.error("Ish boshlanish eslatmasi xatosi: %s", e)


async def handle_staff_video_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """HR guruhiga kelgan krujok (video note) xabarlarini kuzatib,
    yuboruvchini ishga kelgan deb belgilaydi."""
    if not update.effective_chat or update.effective_chat.id != HR_GROUP_ID:
        return
    user = update.effective_user
    if not user or not user.username:
        return
    try:
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")
        username = user.username.lstrip("@")

        state_rows = await sb_get("biznes_data", params={"id": f"eq.{ATTENDANCE_KELGANLAR_ID}"})
        state = state_rows[0]["data"] if state_rows and isinstance(state_rows[0].get("data"), dict) else {}
        kelganlar = state.setdefault(today_str, {})
        if username in kelganlar:
            return  # allaqachon tasdiqlangan
        kelganlar[username] = now.strftime("%H:%M:%S")

        await sb_upsert("biznes_data", ATTENDANCE_KELGANLAR_ID, {"data": state})

        fio = user.full_name or f"@{username}"
        try:
            org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
            org_nodes = org_rows[0]["data"] if org_rows else []
            for n in org_nodes:
                tgs = parse_tg_field(n)
                if username in tgs and n.get("fio"):
                    fio = n["fio"]
                    break
        except Exception:
            pass

        try:
            await context.application.bot.send_message(
                chat_id=HR_GROUP_ID,
                text=f"✅ {fio} — {now.strftime('%H:%M')} da ishga keldi (video orqali tasdiqlandi)",
            )
        except Exception as e:
            log.error("HR guruhga xabar yuborishda xato: %s", e)
        log.info("Davomat qayd etildi: %s -> %s", username, now.strftime("%H:%M:%S"))
    except Exception as e:
        log.error("Davomat qayd etish xatosi: %s", e)


async def check_ish_kelmagan(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har xodimning ish boshlanishidan 30 daqiqa o'tgach, agar krujok
    kelmagan bo'lsa, HR guruhiga kelmaganlar ro'yxatini yuboradi."""
    app = context.application
    try:
        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        employees = _org_employees_with_schedule(org_nodes)
        if not employees:
            return

        today_kun = _today_ish_kun_key()
        now = datetime.now(TASHKENT_TZ)
        today_str = now.strftime("%Y-%m-%d")

        kelganlar_rows = await sb_get("biznes_data", params={"id": f"eq.{ATTENDANCE_KELGANLAR_ID}"})
        kelganlar_state = kelganlar_rows[0]["data"] if kelganlar_rows and isinstance(kelganlar_rows[0].get("data"), dict) else {}
        kelganlar = kelganlar_state.get(today_str, {})
        kelganlar_lower = {u.lower() for u in kelganlar.keys()}

        state_rows = await sb_get("biznes_data", params={"id": f"eq.{ATTENDANCE_KELMAGAN_ID}"})
        state = state_rows[0]["data"] if state_rows and isinstance(state_rows[0].get("data"), dict) else {}
        checked = state.setdefault(today_str, [])

        due_times = set()
        for n in employees:
            kunlar = n.get("ishKunlari")
            if kunlar is not None and today_kun not in kunlar:
                continue
            vaqt, _ = _employee_schedule_for_day(n, today_kun)
            if not vaqt:
                continue
            try:
                h, m = map(int, vaqt.split(":"))
            except Exception:
                continue
            start_dt = now.replace(hour=h, minute=m, second=0, microsecond=0)
            diff_minutes = (now - start_dt).total_seconds() / 60
            if 29 <= diff_minutes <= 30 and vaqt not in checked:
                due_times.add(vaqt)

        if not due_times:
            return

        for vaqt in due_times:
            group = [
                n for n in employees
                if _employee_schedule_for_day(n, today_kun)[0] == vaqt
                and (n.get("ishKunlari") is None or today_kun in n.get("ishKunlari"))
            ]
            missing = []
            for n in group:
                tgs = parse_tg_field(n)
                if not tgs or not any(u.lower() in kelganlar_lower for u in tgs):
                    missing.append(n.get("fio") or n.get("name") or "—")
            checked.append(vaqt)
            if missing:
                lines = [f"❌ Ishga kelmadi ({vaqt} boshlanishi kerak edi, 30 daqiqa o'tdi):"]
                for fio in missing:
                    lines.append(f"👤 {fio}")
                lines.append("")
                lines.append("@umidpulatov")
                await app.bot.send_message(chat_id=HR_GROUP_ID, text="\n".join(lines))
                log.info("Kelmaganlar ro'yxati yuborildi: %s (%d kishi)", vaqt, len(missing))

        await sb_upsert("biznes_data", ATTENDANCE_KELMAGAN_ID, {"data": state})
    except Exception as e:
        log.error("Kelmaganlar tekshiruvi xatosi: %s", e)




# ============================================================
# KAITEN — ERTANGI DEDLAYNLAR HAQIDA KUNLIK ESLATMA (23:00)
# ============================================================
def _kaiten_fmt_deadline(iso: str) -> str:
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00")) if "T" in iso else datetime.fromisoformat(iso)
    except Exception:
        return iso
    return dt.strftime("%d.%m.%Y %H:%M")


async def _kaiten_deadline_digest(context: ContextTypes.DEFAULT_TYPE, days_ahead: int, title: str) -> None:
    """Kaiten vazifalarini (bugun yoki ertaga muddati tugaydiganlarni)
    bo'limlar bo'yicha guruhlab, Sirly xodimlar guruhiga yuboradi.
    Ijrochi va nazoratchi @username orqali taglanadi."""
    app = context.application
    try:
        kaiten_rows = await sb_get("biznes_data", params={"id": "eq.kaiten"})
        kaiten_data = kaiten_rows[0]["data"] if kaiten_rows else {}
        tasks = kaiten_data.get("tasks", []) if isinstance(kaiten_data, dict) else []

        org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
        org_nodes = org_rows[0]["data"] if org_rows else []
        fio_tg_map = build_fio_tg_map(org_nodes)

        target_day = datetime.now(TASHKENT_TZ) + timedelta(days=days_ahead)
        target_str = target_day.strftime("%Y-%m-%d")
        target_display = target_day.strftime("%d.%m.%Y")

        filtered = [
            t for t in tasks
            if not t.get("archived")
            and t.get("status") in ("todo", "progress")
            and t.get("deadline")
            and str(t["deadline"]).split("T")[0] == target_str
        ]

        divider = "━━━━━━━━━━━━━"
        lines = [divider, f"{title} — {target_display}", divider]

        if not filtered:
            lines.append("")
            lines.append(f"✅ ({target_display}) muddati tugaydigan vazifalar yo'q.")
        else:
            by_dept = {}
            for t in filtered:
                dept = t.get("dept") or "Boshqa"
                by_dept.setdefault(dept, []).append(t)

            dept_order = list(KAITEN_DEPTS) + [d for d in by_dept if d not in KAITEN_DEPTS]

            for dept in dept_order:
                dept_tasks = by_dept.get(dept)
                if not dept_tasks:
                    continue
                lines.append("")
                lines.append(f"🏢 {dept}")
                for t in dept_tasks:
                    emp_fio = t.get("empFio") or "—"
                    emp_tg = fio_tg_map.get(norm_fio(emp_fio), "") or (t.get("empTg") or "").strip().lstrip("@")
                    ijrochi_line = emp_fio + (f" @{emp_tg}" if emp_tg else "")

                    naz_fio = t.get("nazFio")
                    if naz_fio:
                        naz_tg = fio_tg_map.get(norm_fio(naz_fio), "")
                        nazoratchi_line = naz_fio + (f" @{naz_tg}" if naz_tg else "")
                    else:
                        nazoratchi_line = "— (belgilanmagan)"

                    lines.append("")
                    lines.append(f'📌 "{t.get("text","")}"')
                    lines.append(f"🙋 Ijrochi: {ijrochi_line}")
                    lines.append(f"👁 Nazoratchi: {nazoratchi_line}")
                    lines.append(f"⏰ Muddat: {_kaiten_fmt_deadline(t.get('deadline',''))}")
                lines.append(divider)

            lines.append(f"Jami: {len(filtered)} ta vazifa muddati ({target_display}) tugaydi.")

        await app.bot.send_message(chat_id=TOPSHIRIQLAR_GROUP_ID, text="\n".join(lines))
        log.info("%s yuborildi: %d ta vazifa", title, len(filtered))
    except Exception as e:
        log.error("Kaiten dedlayn digest xatosi (%s): %s", title, e)


async def daily_deadline_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 23:00 (Toshkent) da ertangi kunga muddati tugaydigan vazifalar."""
    await _kaiten_deadline_digest(context, days_ahead=1, title="⏰ ERTANGI DEDLAYNLAR")


async def daily_today_tasks_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Har kuni 10:00 (Toshkent) da bugun tugaydigan vazifalar ro'yxati."""
    await _kaiten_deadline_digest(context, days_ahead=0, title="📋 BUGUNGI ISHLAR RO'YXATI")


async def check_deadline_2h_before(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Vazifa dedlayniga 2 soat qolganda, Sirly Staff guruhiga bir martalik
    eslatma yuboradi (ijrochi va nazoratchi taglanadi)."""
    app = context.application
    try:
        kaiten_rows = await sb_get("biznes_data", params={"id": "eq.kaiten"})
        if not kaiten_rows:
            return
        kaiten_data = kaiten_rows[0]["data"]
        if not isinstance(kaiten_data, dict):
            return
        tasks = kaiten_data.get("tasks", [])

        now = datetime.now(TASHKENT_TZ)
        due = []
        due_ids = []
        for t in tasks:
            if t.get("archived") or t.get("status") not in ("todo", "progress"):
                continue
            if t.get("deadline2hNotified"):
                continue
            deadline = t.get("deadline")
            if not deadline:
                continue
            try:
                dl = datetime.fromisoformat(deadline)
                if dl.tzinfo is None:
                    dl = dl.replace(tzinfo=TASHKENT_TZ)
            except Exception:
                continue
            diff_minutes = (dl - now).total_seconds() / 60
            if 119 <= diff_minutes <= 121:
                due.append(t)
                due_ids.append(t.get("id"))

        if due:
            org_rows = await sb_get("biznes_data", params={"id": "eq.org"})
            org_nodes = org_rows[0]["data"] if org_rows else []
            fio_tg_map = build_fio_tg_map(org_nodes)

            divider = "━━━━━━━━━━━━━"
            for t in due:
                emp_fio = t.get("empFio") or "—"
                emp_tg = fio_tg_map.get(norm_fio(emp_fio), "") or (t.get("empTg") or "").strip().lstrip("@")
                ijrochi_line = emp_fio + (f" @{emp_tg}" if emp_tg else "")

                naz_fio = t.get("nazFio")
                if naz_fio:
                    naz_tg = fio_tg_map.get(norm_fio(naz_fio), "")
                    nazoratchi_line = naz_fio + (f" @{naz_tg}" if naz_tg else "")
                else:
                    nazoratchi_line = "— (belgilanmagan)"

                lines = [
                    divider,
                    "⏳ DEDLAYNGACHA 2 SOAT QOLDI",
                    divider,
                    "",
                    f'📌 "{t.get("text","")}"',
                    f"🏢 {t.get('dept','—')}",
                    f"🙋 Ijrochi: {ijrochi_line}",
                    f"👁 Nazoratchi: {nazoratchi_line}",
                    f"⏰ Muddat: {_kaiten_fmt_deadline(t.get('deadline',''))}",
                    divider,
                ]
                await app.bot.send_message(chat_id=TOPSHIRIQLAR_GROUP_ID, text="\n".join(lines))
            log.info("2 soatlik dedlayn eslatmasi yuborildi: %d ta vazifa", len(due))

        if due_ids:
            def _mark(fresh_data):
                for ft in fresh_data.get("tasks", []):
                    if ft.get("id") in due_ids:
                        ft["deadline2hNotified"] = True
            await sb_mutate_and_save("kaiten", _mark)
    except Exception as e:
        log.error("2 soatlik dedlayn eslatmasi xatosi: %s", e)


# ============================================================
# DAVRIY TEKSHIRUV (JobQueue)
# ============================================================
async def poll_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    app = context.application
    await process_group_messages(app)
    await process_reminders(app)
    await process_file_requests(app)


# ============================================================
# /test_dedlayn BUYRUG'I — 23:00 ni kutmasdan, xabarni darhol
# yuborib ko'rish uchun (test maqsadida)
# ============================================================
async def test_deadline_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("⏳ Tekshirilmoqda, biroz kuting...")
    await daily_deadline_reminder(context)
    await update.message.reply_text(
        "✅ Tayyor! Sirly xodimlar guruhini tekshiring — xabar o'sha yerga yuborildi."
    )


# ============================================================
# /test_hisobot BUYRUG'I — 18:00 ni kutmasdan, Buxgalteriya
# hisobotlarini (Promokod + Pul berish, matn + Excel fayllar) darhol
# yuborib ko'rish uchun
# ============================================================
async def test_hisobot_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("⏳ Tekshirilmoqda, biroz kuting...")
    await daily_promokod_report(context)
    await weekly_pul_berish_report(context)
    await update.message.reply_text(
        "✅ Tayyor! Buxgalteriya guruhini tekshiring — matn va Excel fayllar o'sha yerga yuborildi."
    )


async def test_baza415_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("⏳ Tekshirilmoqda, biroz kuting...")
    await daily_baza415_report(context)
    await update.message.reply_text(
        "✅ Tayyor! Guruhni tekshiring — 415 baza Excel fayli o'sha yerga yuborildi."
    )


# ============================================================
# /start BUYRUG'I — mini-appni ochish tugmasi
# ============================================================
async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if MINIAPP_URL:
        kb = InlineKeyboardMarkup(
            [[InlineKeyboardButton("📋 Ilovani ochish", web_app=WebAppInfo(url=MINIAPP_URL))]]
        )
        await update.message.reply_text(
            "Xush kelibsiz! Ilovani ochish uchun tugmani bosing:", reply_markup=kb
        )
    else:
        await update.message.reply_text("Bot ishga tushdi. MINIAPP_URL sozlanmagan.")


# ============================================================
# ISHGA TUSHIRISH / TO'XTATISH
# ============================================================
async def post_init(app: Application) -> None:
    global http_client
    http_client = httpx.AsyncClient(timeout=30)
    log.info("Sirly bot ishga tushdi. Poll interval: %s soniya", POLL_SECONDS)


async def post_shutdown(app: Application) -> None:
    if http_client:
        await http_client.aclose()


def main() -> None:
    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .post_init(post_init)
        .post_shutdown(post_shutdown)
        .build()
    )
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("test_dedlayn", test_deadline_cmd))
    app.add_handler(CommandHandler("test_hisobot", test_hisobot_cmd))
    app.add_handler(CommandHandler("test_baza415", test_baza415_cmd))
    app.add_handler(CommandHandler("murojaatlar_hisobot", murojaatlar_hisobot_cmd))
    app.add_handler(CommandHandler("test_jihozlar", test_jihozlar_cmd))
    app.add_handler(CommandHandler("noaktiv_hamkorlar", noaktiv_hamkorlar_cmd))
    app.add_handler(MessageHandler(filters.PHOTO, handle_support_photo))
    app.add_handler(MessageHandler(filters.VIDEO_NOTE, handle_staff_video_note))
    app.add_handler(MessageHandler(filters.VIDEO_NOTE, handle_jihozlar_video), group=1)
    app.job_queue.run_repeating(poll_job, interval=POLL_SECONDS, first=5)

    # Skrinshot so'rovi — har 30 daqiqada, 10:30 dan 23:30 gacha
    for shot_time in generate_screenshot_times():
        app.job_queue.run_daily(screenshot_request_job, time=shot_time)

    # Muammoli mijozlar — davriy TEKSHIRUV xabari O'CHIRILGAN (foydalanuvchi so'rovi bo'yicha)
    # for check_time in generate_check_times():
    #     app.job_queue.run_daily(muammoli_check_job, time=check_time)

    # Muammoli mijozlar — kunlik eslatma, har kuni 10:30, 15:00, 19:00 (Toshkent vaqti)
    app.job_queue.run_daily(daily_muammoli_reminder, time=dt_time(hour=10, minute=30, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_muammoli_reminder, time=dt_time(hour=15, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_muammoli_reminder, time=dt_time(hour=19, minute=0, tzinfo=TASHKENT_TZ))

    # Muammoli mijozlar — 24 soatlik muddat nazorati, har 30 daqiqada tekshiriladi
    app.job_queue.run_repeating(check_overdue_muammoli, interval=1800, first=60)

    # 415 baza — ovqat berish vaqtiga 1 soat qolganda Partnership guruhiga eslatma
    app.job_queue.run_repeating(check_serving_time_reminders, interval=60, first=30)
    app.job_queue.run_repeating(check_calling_status_before_serving, interval=60, first=45)
    app.job_queue.run_repeating(check_zvonok2_call_reminder, interval=60, first=50)

    # Kaiten — ertangi dedlaynlar haqida kunlik eslatma, har kuni 23:00 (Toshkent vaqti)
    app.job_queue.run_daily(daily_deadline_reminder, time=dt_time(hour=23, minute=0, tzinfo=TASHKENT_TZ))

    # Kaiten — bugungi ishlar ro'yxati, har kuni 10:00 (Toshkent vaqti)
    app.job_queue.run_daily(daily_today_tasks_reminder, time=dt_time(hour=10, minute=0, tzinfo=TASHKENT_TZ))

    # Buxgalteriya — haftalik hisobot (Promokod + Pul berish), faqat JUMA 10:00 (Toshkent vaqti)
    app.job_queue.run_daily(daily_promokod_report, time=dt_time(hour=18, minute=0, tzinfo=TASHKENT_TZ), days=(1, 2, 3, 4, 5, 6))
    app.job_queue.run_daily(weekly_pul_berish_report, time=dt_time(hour=18, minute=0, tzinfo=TASHKENT_TZ), days=(5,))
    app.job_queue.run_daily(daily_baza415_report, time=dt_time(hour=10, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_jihozlar_hujjat, time=dt_time(hour=23, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(weekly_jihozlar_hujjat, time=dt_time(hour=9, minute=0, tzinfo=TASHKENT_TZ), days=(6,))
    app.job_queue.run_repeating(check_jihozlar_deadline, interval=60, first=25)
    app.job_queue.run_daily(check_yashil_hamkor_auto, time=dt_time(hour=9, minute=0, tzinfo=TASHKENT_TZ))

    # Ish davomati — ish boshlanishiga 10 daqiqa qolganda eslatma, har daqiqada tekshiriladi
    app.job_queue.run_repeating(check_ish_boshlanish_reminder, interval=60, first=15)
    app.job_queue.run_daily(daily_tomorrow_schedule_reminder, time=dt_time(hour=22, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_it_report_reminder, time=dt_time(hour=22, minute=30, tzinfo=TASHKENT_TZ), days=(1, 2, 3, 4, 5, 6))
    # Ish davomati — ish boshlanishidan 30 daqiqa o'tsa, kelmaganlar ro'yxati
    app.job_queue.run_repeating(check_ish_kelmagan, interval=60, first=20)

    # Partnership — kunlik qo'ng'iroqlar hisoboti, har kuni 23:00 (Toshkent vaqti)
    app.job_queue.run_daily(daily_calling_report, time=dt_time(hour=23, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_zvonok2_report, time=dt_time(hour=23, minute=0, tzinfo=TASHKENT_TZ))
    app.job_queue.run_daily(daily_muammoli_hamkorlar_reminder, time=dt_time(hour=23, minute=0, tzinfo=TASHKENT_TZ))

    # Kaiten — dedlaynga 2 soat qolganda eslatma, har daqiqada tekshiriladi
    app.job_queue.run_repeating(check_deadline_2h_before, interval=60, first=50)

    log.info("Bot polling boshlandi...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
